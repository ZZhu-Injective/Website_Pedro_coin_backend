import os
import asyncio
import discord
from discord.ext import commands
from discord.ui import Button, View, Modal, TextInput
from openpyxl import Workbook, load_workbook
from datetime import datetime
from typing import Dict, Optional, List


class TalentHubBot:
    _instance = None
    
    def __new__(cls):
        if cls._instance is None:
            cls._instance = super().__new__(cls)
            cls._instance.initialized = False
        return cls._instance
    
    def __init__(self):
        if self.initialized:
            return
            
        self.initialized = True
        self.bot = commands.Bot(command_prefix="!", intents=discord.Intents.default())
        self.submission_channel_id = 1374018261578027129
        self.pending_submissions: Dict[str, dict] = {}
        self.pending_updates: Dict[str, dict] = {}
        self.loop = asyncio.new_event_loop()
        self.excel_file = "1.Atalent_submissions.xlsx"
        
        self._ensure_excel_file()
        
        self.bot.event(self.on_ready)
        self.bot.event(self.on_interaction)
    
    def _ensure_excel_file(self):
        """Ensure the Excel file exists with correct structure"""
        if not os.path.exists(self.excel_file):
            self._init_excel_file()
        else:
            try:
                wb = load_workbook(self.excel_file)
                ws = wb.active
                expected_headers = [
                    "Name", "Role", "Injective Role", "Experience", "Education", 
                    "Location", "Availability", "Monthly Rate", "Skills", "Languages",
                    "Discord", "Email", "Phone", "Telegram", "X", "Github",
                    "Wallet Address", "Wallet Type", "NFT Holdings", "Token Holdings",
                    "Portfolio", "CV", "Image url", "Bio", "Submission date", "Status"
                ]
                if [cell.value for cell in ws[1]] != expected_headers:
                    self._backup_and_recreate_excel()
            except Exception as e:
                print(f"Error verifying Database: {e}")
                self._backup_and_recreate_excel()
    
    def _init_excel_file(self):
        """Initialize a new Excel file with headers"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Submissions"
        
        headers = [
            "Name", "Role", "Injective Role", "Experience", "Education", "Location",
            "Availability", "Monthly Rate", "Skills", "Languages", "Discord", "Email",
            "Phone", "Telegram", "X", "Github", "Wallet Address", "Wallet Type",
            "NFT Holdings", "Token Holdings", "Portfolio", "CV", "Image url", "Bio",
            "Submission date", "Status"
        ]
        
        ws.append(headers)
        wb.save(self.excel_file)
    
    def _backup_and_recreate_excel(self):
        """Backup corrupt Excel file and create a new one"""
        backup_name = f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        os.rename(self.excel_file, backup_name)
        self._init_excel_file()
    
    async def _find_submission_row(self, wallet_address: str) -> Optional[int]:
        """Find the row number for a given wallet address"""
        try:
            wb = load_workbook(self.excel_file)
            ws = wb.active
            
            for row in range(2, ws.max_row + 1):
                if str(ws[f'Q{row}'].value).lower() == wallet_address.lower():
                    return row
                    
            return None
            
        except Exception as e:
            print(f"Error searching Database: {e}")
            return None
    
    async def _update_excel_status(self, wallet_address: str, new_status: str) -> bool:
        """Update status in Excel for a specific wallet address"""
        try:
            wb = load_workbook(self.excel_file)
            ws = wb.active
            
            row = await self._find_submission_row(wallet_address)
            if not row:
                return False
            
            ws[f'Z{row}'] = new_status
            wb.save(self.excel_file)
            return True
            
        except Exception as e:
            print(f"Error updating Database status: {e}")
            return False
    
    async def _update_excel_record(self, wallet_address: str, updates: dict, status: str) -> bool:
        """Update multiple fields in Excel for a specific wallet address"""
        try:
            wb = load_workbook(self.excel_file)
            ws = wb.active
            
            row = await self._find_submission_row(wallet_address)
            if not row:
                return False
            
            column_map = {
                'name': 'A',
                'role': 'B',
                'injectiveRole': 'C',
                'experience': 'D',
                'education': 'E',
                'location': 'F',
                'available': 'G',
                'monthlyRate': 'H',
                'skills': 'I',
                'languages': 'J',
                'discord': 'K',
                'email': 'L',
                'phone': 'M',
                'telegram': 'N',
                'X': 'O',
                'github': 'P',
                'walletType': 'R',
                'nftHold': 'S',
                'tokenHold': 'T',
                'portfolio': 'U',
                'cv': 'V',
                'profilePicture': 'W',
                'bio': 'X'
            }
            
            for field, value in updates.items():
                if field in column_map:
                    col = column_map[field]
                    if field == 'available':
                        ws[f'{col}{row}'] = 'Yes' if value else 'No'
                    elif field in ['skills', 'languages']:
                        ws[f'{col}{row}'] = ', '.join(value) if isinstance(value, list) else value
                    else:
                        ws[f'{col}{row}'] = str(value).strip() if value else ''
            
            ws[f'Y{row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            ws[f'Z{row}'] = status
            
            wb.save(self.excel_file)
            return True
            
        except Exception as e:
            print(f"Error updating Database record: {e}")
            return False
    
    async def _get_existing_record(self, wallet_address: str) -> Optional[dict]:
        """Get the existing record for a wallet address"""
        try:
            wb = load_workbook(self.excel_file)
            ws = wb.active
            
            row = await self._find_submission_row(wallet_address)
            if not row:
                return None
            
            record = {
                'name': ws[f'A{row}'].value,
                'role': ws[f'B{row}'].value,
                'injectiveRole': ws[f'C{row}'].value,
                'experience': ws[f'D{row}'].value,
                'education': ws[f'E{row}'].value,
                'location': ws[f'F{row}'].value,
                'available': ws[f'G{row}'].value == 'Yes',
                'monthlyRate': ws[f'H{row}'].value,
                'skills': ws[f'I{row}'].value.split(', ') if ws[f'I{row}'].value else [],
                'languages': ws[f'J{row}'].value.split(', ') if ws[f'J{row}'].value else [],
                'discord': ws[f'K{row}'].value,
                'email': ws[f'L{row}'].value,
                'phone': ws[f'M{row}'].value,
                'telegram': ws[f'N{row}'].value,
                'X': ws[f'O{row}'].value,
                'github': ws[f'P{row}'].value,
                'walletAddress': ws[f'Q{row}'].value,
                'walletType': ws[f'R{row}'].value,
                'nftHold': ws[f'S{row}'].value,
                'tokenHold': ws[f'T{row}'].value,
                'portfolio': ws[f'U{row}'].value,
                'cv': ws[f'V{row}'].value,
                'profilePicture': ws[f'W{row}'].value,
                'bio': ws[f'X{row}'].value,
                'status': ws[f'Z{row}'].value
            }
            
            return record
            
        except Exception as e:
            print(f"Error reading record from Database: {e}")
            return None
    
    async def _save_new_submission(self, data: dict) -> bool:
        """Save a new submission to Excel"""
        try:
            wallet = data.get('walletAddress', '')
            
            row = [
                data.get('name', '').strip(),
                data.get('role', '').strip(),
                data.get('injectiveRole', '').strip(),
                data.get('experience', '').strip(),
                data.get('education', '').strip(),
                data.get('location', '').strip(),
                'Yes' if data.get('available', False) else 'No',
                data.get('monthlyRate', '').strip(),
                ', '.join(data.get('skills', [])),
                ', '.join(data.get('languages', [])),
                data.get('discord', '').strip(),
                data.get('email', '').strip(),
                data.get('phone', '').strip() or '-',
                data.get('telegram', '').strip() or '-',
                data.get('X', '').strip() or '-',
                data.get('github', '').strip() or '-',
                wallet,
                data.get('walletType', '').strip(),
                data.get('nftHold', '').strip(),
                data.get('tokenHold', '').strip(),
                data.get('portfolio', '').strip() or '-',
                data.get('cv', '').strip(),
                data.get('profilePicture', '').strip(),
                data.get('bio', '').strip(),
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                "Pending"
            ]
            
            wb = load_workbook(self.excel_file)
            ws = wb.active
            ws.append(row)
            wb.save(self.excel_file)
            return True
            
        except Exception as e:
            print(f"Error saving submission to Database: {e}")
            return False
    
    async def on_ready(self):
        print(f'✅ Bot logged in as {self.bot.user}')
    
    async def on_interaction(self, interaction: discord.Interaction):
        if interaction.type != discord.InteractionType.component:
            return
            
        try:
            custom_id = interaction.data.get('custom_id', '')
            if ':' not in custom_id:
                return
                
            action_type, action, wallet = custom_id.split(':', 2)
            
            if action_type == "submission":
                await self._handle_submission_interaction(interaction, action, wallet)
            elif action_type == "update":
                await self._handle_update_interaction(interaction, action, wallet)
                
        except ValueError:
            try:
                action, wallet = custom_id.split(':', 1)
                await self._handle_submission_interaction(interaction, action, wallet)
            except:
                return
        except Exception as e:
            print(f"Error handling interaction: {e}")
            await interaction.response.send_message(
                "⚠️ An error occurred while processing your request!",
                ephemeral=True
            )
    
    async def _handle_submission_interaction(self, interaction: discord.Interaction, action: str, wallet: str):
        """Handle interactions for new submissions"""
        submission = self.pending_submissions.get(wallet)
        
        if not submission:
            await interaction.response.send_message(
                "❌ Submission not found or already processed!",
                ephemeral=True
            )
            return
            
        if action == "approve":
            submission['status'] = "Approved"
            success = await self._update_excel_status(wallet, "Approved")
            await self._update_submission_message(interaction, submission)
            
            if success:
                await interaction.response.send_message(
                    "✅ Submission approved and Database updated!",
                    ephemeral=True
                )
            else:
                await interaction.response.send_message(
                    "⚠️ Approved but failed to update Database!",
                    ephemeral=True
                )
                
        elif action == "reject":
            submission['status'] = "Rejected"
            success = await self._update_excel_status(wallet, "Rejected")
            await self._update_submission_message(interaction, submission)
            
            if success:
                await interaction.response.send_message(
                    "❌ Submission rejected and Database updated!",
                    ephemeral=True
                )
            else:
                await interaction.response.send_message(
                    "⚠️ Rejected but failed to update Database!",
                    ephemeral=True
                )
                
        elif action == "changes":
            await interaction.response.send_modal(
                SubmissionChangesModal(submission))
    
    async def _handle_update_interaction(self, interaction: discord.Interaction, action: str, wallet: str):
        """Handle interactions for update requests"""
        update_data = self.pending_updates.get(wallet)
        
        if not update_data:
            await interaction.response.send_message(
                "❌ Update not found or already processed!",
                ephemeral=True
            )
            return
            
        if action == "approve":
            success = await self._update_excel_record(wallet, update_data['updates'], "Approved")
            await self._update_update_message(interaction, update_data, "Approved")
            
            if success:
                await interaction.response.send_message(
                    "✅ Update approved and Database updated!",
                    ephemeral=True
                )
            else:
                await interaction.response.send_message(
                    "⚠️ Approved but failed to update Database!",
                    ephemeral=True
                )
                
        elif action == "reject":
            await self._update_excel_record(wallet, {}, "Rejected")
            await self._update_update_message(interaction, update_data, "Rejected")
            await interaction.response.send_message(
                "❌ Update rejected - Status updated in Database",
                ephemeral=True
            )
                
        elif action == "changes":
            await interaction.response.send_modal(
                UpdateChangesModal(update_data))
    
    async def _update_submission_message(self, interaction: discord.Interaction, submission: dict):
        """Update the Discord message for a submission"""
        try:
            embed = self._create_submission_embed(submission)
            message = await interaction.channel.fetch_message(submission['message_id'])
            await message.edit(embed=embed, view=None)
        except Exception as e:
            print(f"Error updating submission message: {e}")
    
    async def _update_update_message(self, interaction: discord.Interaction, update_data: dict, status: str):
        """Update the Discord message for an update"""
        try:
            embed = self._create_update_embed(update_data, status)
            message = await interaction.channel.fetch_message(update_data['message_id'])
            await message.edit(embed=embed, view=None)
        except Exception as e:
            print(f"Error updating update message: {e}")
    
    def _create_submission_embed(self, data: dict) -> discord.Embed:
        """Create a Discord embed for a talent submission"""
        submission_data = data.get('data', {})
        status = data.get('status', 'Pending')
        
        color_map = {
            "Approved": 0x00ff00,    
            "Rejected": 0xff0000,    
            "Pending": 0xffff00,     
            "Changes Requested": 0xffa500,
            "On Hold": 0x808080       
        }
        color = color_map.get(status, 0xffff00)
        
        embed = discord.Embed(
            title=f"🎯 {submission_data.get('name', 'N/A')} - {status}",
            description=f"**{submission_data.get('role', 'N/A')}** | {submission_data.get('injectiveRole', 'N/A')}",
            color=color
        )
        
        if submission_data.get('profilePicture'):
            embed.set_thumbnail(url=submission_data['profilePicture'])
        
        basic_info = [
            f"**Experience:** {submission_data.get('experience', 'N/A')}",
            f"**Education:** {submission_data.get('education', 'N/A')}",
            f"**Location:** {submission_data.get('location', 'N/A')}",
            f"**Availability:** {'✅' if submission_data.get('available') else '❌'}",
            f"**Rate:** {submission_data.get('monthlyRate', 'N/A')}",
            f"**Wallet:** `{data.get('wallet', 'N/A')[:6]}...{data.get('wallet', 'N/A')[-4:]}`"
        ]
        embed.add_field(name="ℹ️ Basic Info", value="\n".join(basic_info), inline=False)
        
        blockchain_info = [
            f"**NFTs:** {submission_data.get('nftHold', 'N/A')}",
            f"**Tokens:** {submission_data.get('tokenHold', 'N/A')}",
            f"**Wallet Type:** {submission_data.get('walletType', 'N/A')}"
        ]
        embed.add_field(name="🔗 Blockchain Info", value="\n".join(blockchain_info), inline=False)
        
        skills = "• " + "\n• ".join(submission_data.get('skills', [])) if submission_data.get('skills') else "None"
        languages = "• " + "\n• ".join(submission_data.get('languages', [])) if submission_data.get('languages') else "None"
        embed.add_field(name="🛠️ Skills", value=skills, inline=True)
        embed.add_field(name="🗣️ Languages", value=languages, inline=True)
        
        contact_info = [
            f"**Discord:** {submission_data.get('discord', '-')}",
            f"**Email:** {submission_data.get('email', '-')}",
            f"**Phone:** {submission_data.get('phone', '-')}",
            f"**Telegram:** {submission_data.get('telegram', 'N/A') or '-'}",
            f"**X:** {submission_data.get('X', 'N/A') or '-'}",
            f"**GitHub:** {submission_data.get('github', 'N/A') or '-'}"
        ]
        embed.add_field(name="📩 Contact Info", value="\n".join(contact_info), inline=False)
        
        links = []
        if submission_data.get('portfolio'):
            links.append(f"**Portfolio:** {submission_data['portfolio']}")
        if submission_data.get('cv'):
            links.append(f"**CV:** [Download CV]({submission_data['cv']})")
        if links:
            embed.add_field(name="🔗 Links", value="\n".join(links), inline=False)
        
        bio = submission_data.get('bio', 'No bio provided')
        embed.add_field(
            name="📝 Bio", 
            value=f"{bio[:250]}{'...' if len(bio) > 250 else ''}", 
            inline=False
        )
        
        embed.set_footer(text=f"Submitted on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        
        return embed
    
    def _create_update_embed(self, update_data: dict, status: str) -> discord.Embed:
        """Create a Discord embed for a talent update"""
        existing_data = update_data.get('existing_data', {})
        updates = update_data.get('updates', {})
        wallet = update_data.get('wallet', '')
        
        color_map = {
            "Approved": 0x00ff00,    
            "Rejected": 0xff0000,    
            "Pending": 0xffff00,     
            "Changes Requested": 0xffa500
        }
        color = color_map.get(status, 0xffff00)
        
        embed = discord.Embed(
            title=f"🔄 Talent Profile Update - {status}",
            description=f"**{existing_data.get('name', 'N/A')}** | `{wallet[:6]}...{wallet[-4:]}`",
            color=color
        )
        
        if existing_data.get('profilePicture'):
            embed.set_thumbnail(url=existing_data['profilePicture'])
        
        current_field = {"name": "🆕 Proposed Changes", "value": "", "inline": False}
        
        for field, new_value in updates.items():
            old_value = existing_data.get(field, '')
            
            if field == 'available':
                old_value = 'Yes' if old_value else 'No'
                new_value = 'Yes' if new_value else 'No'
            elif field in ['skills', 'languages']:
                old_value = ', '.join(old_value) if isinstance(old_value, list) else old_value
                new_value = ', '.join(new_value) if isinstance(new_value, list) else new_value
            
            entry = (
                f"**{field.capitalize()}:**\n"
                f"`Old:` {old_value}\n"
                f"`New:` {new_value}\n\n"
            )
            
            if len(current_field["value"]) + len(entry) > 1024:
                embed.add_field(**current_field)
                current_field = {
                    "name": "🆕 Proposed Changes (cont.)",
                    "value": entry,
                    "inline": False
                }
            else:
                current_field["value"] += entry
        
        if current_field["value"] or not updates:
            if not updates:
                current_field["value"] = "No specific changes - renewal only"
            embed.add_field(**current_field)
        
        embed.set_footer(text=f"Update requested on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        return embed
    
    def _create_submission_review_buttons(self, wallet: str) -> View:
        """Create the action buttons for submission review"""
        view = View(timeout=None)  
        
        buttons = [
            Button(label="✅ Approve", style=discord.ButtonStyle.success, custom_id=f"submission:approve:{wallet}"),
            Button(label="❌ Reject", style=discord.ButtonStyle.danger, custom_id=f"submission:reject:{wallet}"),
            Button(label="✏️ Request Changes", style=discord.ButtonStyle.primary, custom_id=f"submission:changes:{wallet}")
        ]
        
        for button in buttons:
            view.add_item(button)
            
        return view
    
    def _create_update_review_buttons(self, wallet: str) -> View:
        """Create the action buttons for update review"""
        view = View(timeout=None)  
        
        buttons = [
            Button(label="✅ Approve", style=discord.ButtonStyle.success, custom_id=f"update:approve:{wallet}"),
            Button(label="❌ Reject", style=discord.ButtonStyle.danger, custom_id=f"update:reject:{wallet}"),
        ]
        
        for button in buttons:
            view.add_item(button)
            
        return view
    
    async def post_submission(self, data: dict) -> Optional[discord.Message]:
        """Post a new talent submission to Discord"""
        try:
            channel = self.bot.get_channel(self.submission_channel_id)
            if not channel:
                print("❌ Error: Submission channel not found!")
                return None
                
            wallet = data['walletAddress']
            submission = {
                'data': data,
                'status': "Pending",
                'wallet': wallet
            }
            
            excel_success = await self._save_new_submission(data)
            if not excel_success:
                print(f"⚠️ Warning: Failed to save submission for {wallet} to Database")
                
            self.pending_submissions[wallet] = submission
            
            embed = self._create_submission_embed(submission)
            view = self._create_submission_review_buttons(wallet)
            
            message = await channel.send(embed=embed, view=view)
            submission['message_id'] = message.id
            
            return message
            
        except Exception as e:
            print(f"❌ Error posting submission: {e}")
            return None
    
    async def post_update_request(self, wallet_address: str, updates: dict) -> Optional[discord.Message]:
        """Post a talent profile update request to Discord"""
        try:
            channel = self.bot.get_channel(self.submission_channel_id)
            if not channel:
                print("❌ Error: Submission channel not found!")
                return None
                
            existing_data = await self._get_existing_record(wallet_address)
            if not existing_data:
                print(f"❌ Error: No existing record found for wallet {wallet_address}")
                return None
                
            update_data = {
                'existing_data': existing_data,
                'updates': updates,
                'wallet': wallet_address,
                'status': "Pending"
            }
            
            self.pending_updates[wallet_address] = update_data
            
            embed = self._create_update_embed(update_data, "Pending")
            view = self._create_update_review_buttons(wallet_address)
            
            message = await channel.send(embed=embed, view=view)
            update_data['message_id'] = message.id
            
            return message
            
        except Exception as e:
            print(f"❌ Error posting update request: {e}")
            return None
    
    def start(self):
        """Start the Discord bot"""
        async def runner():
            try:
                await self.bot.start("")
            except Exception as e:
                print(f"❌ Bot error: {e}")
                await self.bot.close()
        
        self.loop.run_until_complete(runner())
    
    def stop(self):
        """Stop the Discord bot gracefully"""
        async def shutdown():
            await self.bot.close()
        
        self.loop.run_until_complete(shutdown())
        self.loop.close()


class SubmissionChangesModal(Modal, title="Edit Submission"):
    def __init__(self, submission: dict):
        super().__init__()
        self.submission = submission
        submission_data = submission['data']
        
        self.name = TextInput(
            label="Name",
            default=submission_data.get('name', ''),
            required=True,
            style=discord.TextStyle.short
        )
        self.add_item(self.name)
        
        self.role = TextInput(
            label="Role",
            default=submission_data.get('role', ''),
            required=True,
            style=discord.TextStyle.short
        )
        self.add_item(self.role)
        
        self.injective_role = TextInput(
            label="Injective Role",
            default=submission_data.get('injectiveRole', ''),
            required=True,
            style=discord.TextStyle.short
        )
        self.add_item(self.injective_role)
        
        self.experience = TextInput(
            label="Experience",
            default=submission_data.get('experience', ''),
            required=True,
            style=discord.TextStyle.short
        )
        self.add_item(self.experience)
        
        self.education = TextInput(
            label="Education",
            default=submission_data.get('education', ''),
            required=True,
            style=discord.TextStyle.short
        )
        self.add_item(self.education)
        
    async def on_submit(self, interaction: discord.Interaction):
        submission_data = self.submission['data']
        
        submission_data.update({
            'name': self.name.value,
            'role': self.role.value,
            'injectiveRole': self.injective_role.value,
            'experience': self.experience.value,
            'education': self.education.value
        })
        
        self.submission['status'] = "Approved"
        wallet = self.submission['wallet']
        success = await TalentHubBot()._update_excel_status(wallet, "Approved")
        
        await TalentHubBot()._update_submission_message(interaction, self.submission)
        
        if success:
            await interaction.response.send_message(
                "✅ Changes saved and submission approved! Database updated.",
                ephemeral=True
            )
        else:
            await interaction.response.send_message(
                "⚠️ Changes saved and approved but failed to update Database!",
                ephemeral=True
            )


class UpdateChangesModal(Modal, title="Edit Update Request"):
    def __init__(self, update_data: dict):
        super().__init__()
        self.update_data = update_data
        
        self.notes = TextInput(
            label="Changes Requested",
            placeholder="Specify what changes are needed...",
            required=True,
            style=discord.TextStyle.long
        )
        self.add_item(self.notes)
    
    async def on_submit(self, interaction: discord.Interaction):
        success = await TalentHubBot()._update_excel_record(
            self.update_data['wallet'], 
            {}, 
            "Changes Requested"
        )
        
        wallet = self.update_data['wallet']
        self.update_data['status'] = "Changes Requested"
        
        await TalentHubBot()._update_update_message(interaction, self.update_data, "Changes Requested")
        
        try:
            submitter = await interaction.guild.fetch_member(int(wallet)) 
            if submitter:
                await submitter.send(
                    f"ℹ️ Your talent profile update for wallet `{wallet[:6]}...{wallet[-4:]}` requires changes:\n"
                    f"```{self.notes.value}```\n"
                    "Please submit a new update with the requested changes."
                )
        except Exception as e:
            print(f"Couldn't notify submitter: {e}")
        
        if success:
            await interaction.response.send_message(
                "✏️ Changes requested and status updated in Database!",
                ephemeral=True
            )
        else:
            await interaction.response.send_message(
                "⚠️ Changes requested but failed to update Database status!",
                ephemeral=True
            )

talent_hub_bot = TalentHubBot()