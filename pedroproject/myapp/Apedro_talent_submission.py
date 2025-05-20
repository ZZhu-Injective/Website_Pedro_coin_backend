import asyncio
import discord
from discord.ext import commands
from discord.ui import Button, View, Modal, TextInput
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
from typing import Dict, Optional, List

class DiscordBot:
    _instance = None
    
    def __new__(cls):
        if cls._instance is None:
            cls._instance = super().__new__(cls)
            cls._instance.initialized = False
        return cls._instance
    
    def __init__(self):
        if self.initialized:
            return
            
        self.initialized = True
        self.bot = commands.Bot(command_prefix="!", intents=discord.Intents.default())
        self.submission_channel_id = 1374018261578027129
        self.pending_submissions: Dict[str, dict] = {}
        self.loop = asyncio.new_event_loop()
        self.excel_file = "1.Atalent_submissions.xlsx"
        
        self._ensure_excel_file()
        
        self.bot.event(self.on_ready)
        self.bot.event(self.on_interaction)
    
    def _ensure_excel_file(self):
        """Ensure the Excel file exists with correct structure"""
        if not os.path.exists(self.excel_file):
            self._init_excel_file()
        else:
            try:
                wb = load_workbook(self.excel_file)
                ws = wb.active
                expected_headers = [
                    "Name", "Role", "Injective Role", "Experience", "Education", 
                    "Location", "Availability", "Monthly Rate", "Skills", "Languages",
                    "Discord", "Email", "Phone", "Telegram", "LinkedIn", "Github",
                    "Wallet Address", "Wallet Type", "NFT Holdings", "Token Holdings",
                    "Portfolio", "CV", "Image url", "Bio", "Submission date", "Status"
                ]
                if [cell.value for cell in ws[1]] != expected_headers:
                    self._backup_and_recreate_excel()
            except Exception as e:
                print(f"Error verifying Database: {e}")
                self._backup_and_recreate_excel()
    
    def _init_excel_file(self):
        """Initialize a new Excel file with headers"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Submissions"
        
        headers = [
            "Name", "Role", "Injective Role", "Experience", "Education", "Location",
            "Availability", "Monthly Rate", "Skills", "Languages", "Discord", "Email",
            "Phone", "Telegram", "LinkedIn", "Github", "Wallet Address", "Wallet Type",
            "NFT Holdings", "Token Holdings", "Portfolio", "CV", "Image url", "Bio",
            "Submission date", "Status"
        ]
        
        ws.append(headers)
        wb.save(self.excel_file)
    
    def _backup_and_recreate_excel(self):
        """Backup corrupt Excel file and create a new one"""
        backup_name = f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        os.rename(self.excel_file, backup_name)
        self._init_excel_file()
    
    async def _update_excel_status(self, wallet_address: str, new_status: str) -> bool:
        """Update status in Excel for a specific wallet address"""
        try:
            wb = load_workbook(self.excel_file)
            ws = wb.active
            
            for row in range(2, ws.max_row + 1):
                if str(ws[f'Q{row}'].value) == wallet_address:
                    ws[f'Z{row}'] = new_status
                    wb.save(self.excel_file)
                    return True
            
            return False
            
        except Exception as e:
            print(f"Error updating Database status: {e}")
            return False
    
    async def _save_new_submission(self, data: dict) -> bool:
        """Save a new submission to Excel"""
        try:
            wallet = data.get('walletAddress', '')
            
            row = [
                data.get('name', '').strip(),
                data.get('role', '').strip(),
                data.get('injectiveRole', '').strip(),
                data.get('experience', '').strip(),
                data.get('education', '').strip(),
                data.get('location', '').strip(),
                'Yes' if data.get('available', False) else 'No',
                data.get('monthlyRate', '').strip(),
                ', '.join(data.get('skills', [])),
                ', '.join(data.get('languages', [])),
                data.get('discord', '').strip(),
                data.get('email', '').strip(),
                data.get('phone', '').strip(),
                data.get('telegram', '').strip() or '-',
                data.get('linkedin', '').strip() or '-',
                data.get('github', '').strip() or '-',
                wallet,
                data.get('walletType', '').strip(),
                data.get('nftHold', '').strip(),
                data.get('tokenHold', '').strip(),
                data.get('portfolio', '').strip() or '-',
                data.get('cv', '').strip(),
                data.get('profilePicture', '').strip(),
                data.get('bio', '').strip(),
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                "Pending"
            ]
            
            wb = load_workbook(self.excel_file)
            ws = wb.active
            ws.append(row)
            wb.save(self.excel_file)
            return True
            
        except Exception as e:
            print(f"Error saving submission to Database: {e}")
            return False
    
    async def on_ready(self):
        print(f'✅ Bot logged in as {self.bot.user}')
        
    async def on_interaction(self, interaction: discord.Interaction):
        if interaction.type != discord.InteractionType.component:
            return
            
        try:
            custom_id = interaction.data.get('custom_id', '')
            if ':' not in custom_id:
                return
                
            action, wallet = custom_id.split(':', 1)
            submission = self.pending_submissions.get(wallet)
            
            if not submission:
                await interaction.response.send_message(
                    "❌ Submission not found or already processed!",
                    ephemeral=True
                )
                return
                
            if action == "approve":
                submission['status'] = "Approved"
                success = await self._update_excel_status(wallet, "Approved")
                await self._update_submission_message(interaction, submission)
                
                if success:
                    await interaction.response.send_message(
                        "✅ Submission approved and Database updated!",
                        ephemeral=True
                    )
                else:
                    await interaction.response.send_message(
                        "⚠️ Approved but failed to update Database!",
                        ephemeral=True
                    )
                    
            elif action == "reject":
                submission['status'] = "Rejected"
                success = await self._update_excel_status(wallet, "Rejected")
                await self._update_submission_message(interaction, submission)
                
                if success:
                    await interaction.response.send_message(
                        "❌ Submission rejected and Database updated!",
                        ephemeral=True
                    )
                else:
                    await interaction.response.send_message(
                        "⚠️ Rejected but failed to update Database!",
                        ephemeral=True
                    )
                    
            elif action == "changes":
                await interaction.response.send_modal(
                    ChangesRequestModal(submission)
                )
                
        except Exception as e:
            print(f"Error handling interaction: {e}")
            await interaction.response.send_message(
                "⚠️ An error occurred while processing your request!",
                ephemeral=True
            )
    
    async def _update_submission_message(self, interaction: discord.Interaction, submission: dict):
        """Update the Discord message for a submission"""
        try:
            embed = self._create_submission_embed(submission)
            message = await interaction.channel.fetch_message(submission['message_id'])
            await message.edit(embed=embed, view=None)
        except Exception as e:
            print(f"Error updating submission message: {e}")
    
    def _create_submission_embed(self, data: dict) -> discord.Embed:
        """Create a Discord embed for a talent submission"""
        submission_data = data.get('data', {})
        status = data.get('status', 'Pending')
        
        color_map = {
            "Approved": 0x00ff00,    
            "Rejected": 0xff0000,    
            "Pending": 0xffff00,     
            "Changes Requested": 0xffa500,
            "On Hold": 0x808080       
        }
        color = color_map.get(status, 0xffff00)
        
        embed = discord.Embed(
            title=f"🎯 {submission_data.get('name', 'N/A')} - {status}",
            description=f"**{submission_data.get('role', 'N/A')}** | {submission_data.get('injectiveRole', 'N/A')}",
            color=color
        )
        
        if submission_data.get('profilePicture'):
            embed.set_thumbnail(url=submission_data['profilePicture'])
        
        basic_info = [
            f"**Experience:** {submission_data.get('experience', 'N/A')}",
            f"**Education:** {submission_data.get('education', 'N/A')}",
            f"**Location:** {submission_data.get('location', 'N/A')}",
            f"**Availability:** {'✅' if submission_data.get('available') else '❌'}",
            f"**Rate:** {submission_data.get('monthlyRate', 'N/A')}",
            f"**Wallet:** `{data.get('wallet', 'N/A')[:6]}...{data.get('wallet', 'N/A')[-4:]}`"
        ]
        embed.add_field(name="ℹ️ Basic Info", value="\n".join(basic_info), inline=False)
        
        blockchain_info = [
            f"**NFTs:** {submission_data.get('nftHold', 'N/A')}",
            f"**Tokens:** {submission_data.get('tokenHold', 'N/A')}",
            f"**Wallet Type:** {submission_data.get('walletType', 'N/A')}"
        ]
        embed.add_field(name="🔗 Blockchain Info", value="\n".join(blockchain_info), inline=False)
        
        skills = "• " + "\n• ".join(submission_data.get('skills', [])) if submission_data.get('skills') else "None"
        languages = "• " + "\n• ".join(submission_data.get('languages', [])) if submission_data.get('languages') else "None"
        embed.add_field(name="🛠️ Skills", value=skills, inline=True)
        embed.add_field(name="🗣️ Languages", value=languages, inline=True)
        
        contact_info = [
            f"**Discord:** {submission_data.get('discord', 'N/A')}",
            f"**Email:** {submission_data.get('email', 'N/A')}",
            f"**Phone:** {submission_data.get('phone', 'N/A')}",
            f"**Telegram:** {submission_data.get('telegram', 'N/A') or '-'}",
            f"**LinkedIn:** {submission_data.get('linkedin', 'N/A') or '-'}",
            f"**GitHub:** {submission_data.get('github', 'N/A') or '-'}"
        ]
        embed.add_field(name="📩 Contact Info", value="\n".join(contact_info), inline=False)
        
        links = []
        if submission_data.get('portfolio'):
            links.append(f"**Portfolio:** {submission_data['portfolio']}")
        if submission_data.get('cv'):
            links.append(f"**CV:** [Download CV]({submission_data['cv']})")
        if links:
            embed.add_field(name="🔗 Links", value="\n".join(links), inline=False)
        
        bio = submission_data.get('bio', 'No bio provided')
        embed.add_field(
            name="📝 Bio", 
            value=f"{bio[:250]}{'...' if len(bio) > 250 else ''}", 
            inline=False
        )
        
        embed.set_footer(text=f"Submitted on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        
        return embed
    
    async def post_submission(self, data: dict) -> Optional[discord.Message]:
        """Post a new talent submission to Discord"""
        try:
            channel = self.bot.get_channel(self.submission_channel_id)
            if not channel:
                print("❌ Error: Submission channel not found!")
                return None
                
            wallet = data['walletAddress']
            submission = {
                'data': data,
                'status': "Pending",
                'wallet': wallet
            }
            
            excel_success = await self._save_new_submission(data)
            if not excel_success:
                print(f"⚠️ Warning: Failed to save submission for {wallet} to Database")
                
            self.pending_submissions[wallet] = submission
            
            embed = self._create_submission_embed(submission)
            view = self._create_review_buttons(wallet)
            
            message = await channel.send(embed=embed, view=view)
            submission['message_id'] = message.id
            
            return message
            
        except Exception as e:
            print(f"❌ Error posting submission: {e}")
            return None
    
    def _create_review_buttons(self, wallet: str) -> View:
        """Create the action buttons for submission review"""
        view = View(timeout=None)  
        
        buttons = [
            Button(label="✅ Approve", style=discord.ButtonStyle.success, custom_id=f"approve:{wallet}"),
            Button(label="❌ Reject", style=discord.ButtonStyle.danger, custom_id=f"reject:{wallet}"),
            Button(label="✏️ Request Changes", style=discord.ButtonStyle.primary, custom_id=f"changes:{wallet}")
        ]
        
        for button in buttons:
            view.add_item(button)
            
        return view
    
    def start(self):
        """Start the Discord bot"""
        async def runner():
            try:
                await self.bot.start('')
            except Exception as e:
                print(f"❌ Bot error: {e}")
                await self.bot.close()
        
        self.loop.run_until_complete(runner())
    
    def stop(self):
        """Stop the Discord bot gracefully"""
        async def shutdown():
            await self.bot.close()
        
        self.loop.run_until_complete(shutdown())
        self.loop.close()

class ChangesRequestModal(Modal, title="Edit Submission"):
    def __init__(self, submission: dict):
        super().__init__()
        self.submission = submission
        submission_data = submission['data']
        
        self.name = TextInput(
            label="Name",
            default=submission_data.get('name', ''),
            required=True,
            style=discord.TextStyle.short
        )
        self.add_item(self.name)
        
        self.role = TextInput(
            label="Role",
            default=submission_data.get('role', ''),
            required=True,
            style=discord.TextStyle.short
        )
        self.add_item(self.role)
        
        self.injective_role = TextInput(
            label="Injective Role",
            default=submission_data.get('injectiveRole', ''),
            required=True,
            style=discord.TextStyle.short
        )
        self.add_item(self.injective_role)
        
        self.experience = TextInput(
            label="Experience",
            default=submission_data.get('experience', ''),
            required=True,
            style=discord.TextStyle.short
        )
        self.add_item(self.experience)
        
        self.education = TextInput(
            label="Education",
            default=submission_data.get('education', ''),
            required=True,
            style=discord.TextStyle.short
        )
        self.add_item(self.education)
        
        
    async def on_submit(self, interaction: discord.Interaction):
        submission_data = self.submission['data']
        
        submission_data.update({
            'name': self.name.value,
            'role': self.role.value,
            'injectiveRole': self.injective_role.value,
            'experience': self.experience.value,
            'education': self.education.value
        })
        
        self.submission['status'] = "Approved"
        wallet = self.submission['wallet']
        success = await DiscordBot()._update_excel_status(wallet, "Approved")
        
        await DiscordBot()._update_submission_message(interaction, self.submission)
        
        if success:
            await interaction.response.send_message(
                "✅ Changes saved and submission approved! Database updated.",
                ephemeral=True
            )
        else:
            await interaction.response.send_message(
                "⚠️ Changes saved and approved but failed to update Database!",
                ephemeral=True
            )

discord_bot = DiscordBot()