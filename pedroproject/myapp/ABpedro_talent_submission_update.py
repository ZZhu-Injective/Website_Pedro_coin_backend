import os
import asyncio
import discord
from discord.ext import commands
from discord.ui import Button, View, Select
from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta
from typing import Dict, Optional, List, Tuple, Any
from dotenv import load_dotenv
import traceback
import queue as thread_queue
import threading

load_dotenv()

class TalentHubBot:
    _instance = None
    _bot_started = False
    
    def __new__(cls):
        if cls._instance is None:
            cls._instance = super().__new__(cls)
            cls._instance.initialized = False
        return cls._instance
    
    def __init__(self):
        if self.initialized:
            return
            
        self.initialized = True
        self.excel_file = "Atalent_submissions.xlsx"
        self.bot_code = os.getenv("DISCORD_BOT")
        
        if self.bot_code:
            intents = discord.Intents.default()
            intents.message_content = True
            
            self.bot = commands.Bot(
                command_prefix="!", 
                intents=intents,
                max_messages=None
            )
            
            self._bot_running = False
            self._bot_loop = None
            self._bot_ready = threading.Event()
            
            self._setup_bot()
        else:
            print("⚠️ DISCORD_BOT token not found. Bot functionality disabled.")
            self.bot = None
        
        self.submission_channel_id = 1374018261578027129
        
        self._lock = asyncio.Lock()
        self._thread_submission_queue = thread_queue.Queue()
        self._queue_processing = False
        
        self._ensure_excel_file()
            
    def _setup_bot(self):
        """Setup bot events and commands"""
        @self.bot.event
        async def on_ready():
            await self._on_ready_callback()
        
        self.bot.tree.command(name="job_open")(self.job_open_command)
        self.bot.tree.command(name="job_show")(self.show_command)
        self.bot.tree.command(name="job_change")(self.change_command)
        self.bot.tree.command(name="job_remove")(self.remove_command)
        self.bot.tree.command(name="job_variable")(self.column_names_command)
        self.bot.tree.command(name="job_status")(self.job_status_command)
        
        @self.bot.command(name="job_open")
        async def job_open_prefix(ctx):
            await self.job_open_prefix_command(ctx)

        @self.bot.command(name="job_status")
        async def job_status_prefix(ctx):
            await self.job_status_prefix_command(ctx)
            
        @self.bot.command(name="job_show")
        async def show_prefix(ctx):
            await self.show_prefix_command(ctx)
            
        @self.bot.command(name="job_change")
        async def change_prefix(ctx, wallet_address: str, column_name: str, *, new_value: str):
            await self.change_prefix_command(ctx, wallet_address, column_name, new_value)
            
        @self.bot.command(name="job_remove")
        async def remove_prefix(ctx, wallet_address: str):
            await self.remove_prefix_command(ctx, wallet_address)
            
        @self.bot.command(name="job_variable")
        async def column_names_prefix(ctx):
            await self.column_names_prefix_command(ctx)
            
        @self.bot.event
        async def on_interaction(interaction):
            await self._handle_interaction(interaction)
    
    async def _on_ready_callback(self):
        """Callback when bot is ready"""
        print(f'✅ Bot logged in as {self.bot.user}')
        print(f'✅ Bot ID: {self.bot.user.id}')
        print(f'✅ Submission channel ID: {self.submission_channel_id}')
        
        self._bot_loop = asyncio.get_event_loop()
        self._bot_ready.set()
        self._bot_running = True
        print(f"✅ Bot ready and loop set: {self._bot_loop}")
        
        await self._test_excel_access()
        
        channel = self.bot.get_channel(self.submission_channel_id)
        if channel:
            print(f'✅ Found submission channel: #{channel.name} ({channel.id})')
            permissions = channel.permissions_for(channel.guild.me)
            print(f'✅ Bot permissions in channel:')
            print(f'   - Send Messages: {permissions.send_messages}')
            print(f'   - Embed Links: {permissions.embed_links}')
            print(f'   - Read Messages: {permissions.read_messages}')
        else:
            print(f"❌ Cannot find channel with ID {self.submission_channel_id}")
            print('💡 Make sure:')
            print('   1. The bot is added to your Discord server')
            print('   2. The channel ID is correct')
            print('   3. The bot has permission to view the channel')
        
        try:
            synced = await self.bot.tree.sync()
            print(f"✅ Synced {len(synced)} command(s)")
            
            counts = await self._get_status_counts()
            
            if channel:
                notification = f"🤖 **Talent Hub Bot is ready!**\n"
                notification += f"📊 **Status:** {counts['Pending']} pending, {counts['Approved']} approved, {counts['Rejected']} rejected\n"
                notification += f"💡 Use `/job_open` to review pending submissions"
                
                ready_embed = discord.Embed(
                    title="🤖 Talent Hub Bot Online",
                    description=notification,
                    color=discord.Color.green()
                )
                ready_embed.set_footer(text="Ready to receive new submissions!")
                await channel.send(embed=ready_embed)
            
            # Start processing queued submissions
            asyncio.create_task(self._process_queued_submissions())
            
            # Start periodic notifications
            asyncio.create_task(self._periodic_notifications())
            
        except Exception as e:
            print(f"❌ Error during bot ready: {e}")
            traceback.print_exc()
    
    async def _process_queued_submissions(self):
        """Process any submissions that were queued before bot was ready"""
        print("🔄 Processing queued submissions...")
        
        # Wait a moment for everything to stabilize
        await asyncio.sleep(5)
        
        try:
            while not self._thread_submission_queue.empty():
                try:
                    data = self._thread_submission_queue.get_nowait()
                    print(f"🔄 Processing queued submission for: {data.get('walletAddress', 'Unknown')}")
                    await self._process_new_submission(data)
                    self._thread_submission_queue.task_done()
                except Exception as e:
                    print(f"❌ Error processing queued submission: {e}")
                    traceback.print_exc()
            
            print("✅ All queued submissions processed")
            
        except Exception as e:
            print(f"❌ Error in queue processing: {e}")
            traceback.print_exc()
    
    async def _handle_interaction(self, interaction: discord.Interaction):
        """Handle button interactions"""
        if interaction.type != discord.InteractionType.component:
            return
            
        try:
            custom_id = interaction.data.get('custom_id', '')
            
            if ':' not in custom_id:
                return
                
            parts = custom_id.split(':', 2)
            
            if len(parts) == 2:
                # New format for quick actions: "quick_approve:wallet" or "review:wallet"
                action, wallet = parts
                
                print(f"[QUICK ACTION] Action: {action}, Wallet: {wallet}")
                
                # Respond immediately to avoid timeout
                if not interaction.response.is_done():
                    await interaction.response.defer(ephemeral=True)
                
                if action == "quick_approve":
                    await self._handle_quick_approve(interaction, wallet)
                elif action == "review":
                    await self._handle_review_button(interaction, wallet)
                else:
                    if not interaction.response.is_done():
                        await interaction.response.send_message(
                            f"❌ Unknown action: {action}",
                            ephemeral=True
                        )
            
            elif len(parts) == 3:
                action_type, action, wallet = parts
                
                # Debug logging
                print(f"[INTERACTION] Type: {action_type}, Action: {action}, Wallet: {wallet}")
                
                # Respond immediately to avoid timeout
                if not interaction.response.is_done():
                    await interaction.response.defer(ephemeral=True)
                
                if action_type == "submission":
                    print(f"[INTERACTION] Routing to submission handler")
                    asyncio.create_task(self._handle_submission_interaction(interaction, action, wallet))
                else:
                    print(f"[INTERACTION WARNING] Unknown action type: {action_type}")
                    if not interaction.response.is_done():
                        await interaction.response.send_message(
                            f"❌ Unknown action type: {action_type}",
                            ephemeral=True
                        )
            else:
                print(f"[INTERACTION ERROR] Invalid parts count: {len(parts)}")
                if not interaction.response.is_done():
                    await interaction.response.send_message(
                        "❌ Invalid interaction format!",
                        ephemeral=True
                    )
                
        except Exception as e:
            print(f"[INTERACTION ERROR] Error handling interaction: {e}")
            traceback.print_exc()
            try:
                if not interaction.response.is_done():
                    await interaction.response.send_message(
                        "❌ An error occurred while processing your request!",
                        ephemeral=True
                    )
            except:
                pass
    
    async def _handle_quick_approve(self, interaction: discord.Interaction, wallet: str):
        """Handle quick approve from notification"""
        try:
            # Update status in Excel
            success = await self._update_excel_status(wallet, "Approved")
            
            if success:
                # Get record details for confirmation
                record = await self._get_record_details(wallet)
                
                if record:
                    name = record.get('Name', 'Unknown')
                    # Update the original notification message
                    try:
                        # Create updated embed
                        updated_embed = discord.Embed(
                            title="✅ **SUBMISSION APPROVED** ✅",
                            description=f"**{name}** has been approved by {interaction.user.mention}",
                            color=discord.Color.green()
                        )
                        updated_embed.add_field(name="👤 Name", value=name, inline=True)
                        updated_embed.add_field(name="💼 Role", value=record.get('Role', 'N/A'), inline=True)
                        updated_embed.add_field(name="💰 Wallet", value=f"`{wallet[:8]}...{wallet[-6:]}`", inline=True)
                        updated_embed.add_field(name="📅 Approved At", 
                                              value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'), 
                                              inline=False)
                        updated_embed.set_footer(text="Use /job_show to view all approved submissions")
                        
                        # Remove buttons from original message
                        await interaction.message.edit(embed=updated_embed, view=None)
                    except Exception as e:
                        print(f"Error updating message: {e}")
                
                await interaction.followup.send(
                    f"✅ **Quick Approved:** {name}'s submission has been approved!",
                    ephemeral=True
                )
                
                # Send confirmation to main channel
                channel = self.bot.get_channel(self.submission_channel_id)
                if channel:
                    confirmation_embed = discord.Embed(
                        title="✅ Submission Approved via Quick Action",
                        description=f"**{name}** was approved by {interaction.user.mention}",
                        color=discord.Color.green()
                    )
                    confirmation_embed.add_field(name="Wallet", value=f"`{wallet}`", inline=False)
                    await channel.send(embed=confirmation_embed)
            else:
                await interaction.followup.send(
                    f"❌ Failed to approve submission for wallet `{wallet}`",
                    ephemeral=True
                )
                
        except Exception as e:
            print(f"[QUICK APPROVE ERROR] {e}")
            traceback.print_exc()
            await interaction.followup.send(
                "❌ An error occurred while processing quick approve.",
                ephemeral=True
            )

    async def job_open_command(self, interaction: discord.Interaction):
        """
        Open pending submissions to review.
        Shows dropdown with all pending submissions (name and role).
        When selected, shows embed with details and Approve/Reject/Close buttons.
        Updates Excel status accordingly.
        """
        await interaction.response.defer()
        
        try:
            pending_records = await self._get_pending_records()
            
            if not pending_records:
                embed = discord.Embed(
                    title="📭 Pending Submissions",
                    description="No submissions waiting for review.",
                    color=discord.Color.green()
                )
                await interaction.followup.send(embed=embed)
                return
            
            embed = discord.Embed(
                title="📬 Pending Submissions",
                description=f"Select a submission to review:",
                color=discord.Color.blue()
            )
            embed.add_field(name="📋 Total Pending", value=f"{len(pending_records)} submission(s)", inline=False)
            embed.set_footer(text="Select a submission from the dropdown below")
            
            original_message = None
            
            class SubmissionSelect(discord.ui.Select):
                def __init__(self, bot_instance, records):
                    options = []
                    for name, wallet in records[:25]: 
                        options.append(discord.SelectOption(
                            label=f"{name[:90]}",
                            description=f"Click to review",
                            value=wallet,
                            emoji="📄"
                        ))
                    
                    super().__init__(
                        placeholder="Choose a submission to review...",
                        min_values=1,
                        max_values=1,
                        options=options
                    )
                    self.bot_instance = bot_instance
                    self.records = records 
                
                async def callback(self, interaction: discord.Interaction):
                    await interaction.response.defer()
                    wallet = self.values[0]
                    
                    record = await self.bot_instance._get_record_details(wallet)
                    
                    if not record:
                        await interaction.followup.send(
                            f"❌ Could not find details for wallet: `{wallet}`",
                            ephemeral=True
                        )
                        return
                    
                    detail_embed = discord.Embed(
                        title=f"📄 Submission Review",
                        description=f"**{record.get('Name', 'N/A')}** - {record.get('Role', 'N/A')}",
                        color=discord.Color.blue()
                    )
                    
                    detail_embed.add_field(
                        name="Basic Info",
                        value=f"**Experience:** {record.get('Experience', 'N/A')}\n"
                            f"**Education:** {record.get('Education', 'N/A')}\n"
                            f"**Location:** {record.get('Location', 'N/A')}\n"
                            f"**Available:** {record.get('Availability', 'N/A')}",
                        inline=False
                    )
                    
                    detail_embed.add_field(
                        name="Wallet",
                        value=f"`{wallet}`",
                        inline=False
                    )
                    
                    skills = record.get('Skills', 'N/A')
                    detail_embed.add_field(
                        name="Skills",
                        value=skills[:500] + "..." if len(skills) > 500 else skills,
                        inline=False
                    )
                    
                    languages = record.get('Languages', 'N/A')
                    detail_embed.add_field(
                        name="Languages",
                        value=languages[:200] + "..." if len(languages) > 200 else languages,
                        inline=False
                    )
                    
                    detail_embed.set_footer(text=f"Wallet: {wallet[:8]}...{wallet[-6:]}")
                    
                    class ReviewView(discord.ui.View):
                        def __init__(self, bot_instance, wallet, user_name, original_view):
                            super().__init__(timeout=180)
                            self.bot_instance = bot_instance
                            self.wallet = wallet
                            self.user_name = user_name
                            self.original_view = original_view
                        
                        @discord.ui.button(label="Approve", style=discord.ButtonStyle.success, emoji="✅", row=0)
                        async def approve_button(self, interaction: discord.Interaction, button: discord.ui.Button):
                            await interaction.response.defer()
                            
                            success = await self.bot_instance._update_excel_status(self.wallet, "Approved")
                            
                            if success:
                                confirmation_embed = discord.Embed(
                                    title="✅ Submission Approved",
                                    description=f"**{self.wallet}** has been approved by {self.user_name}.",
                                    color=discord.Color.green()
                                )
                                confirmation_embed.add_field(name="Status", value="✅ **APPROVED**", inline=True)
                                confirmation_embed.add_field(name="Action", value="✅ Approved", inline=True)
                                confirmation_embed.set_footer(text=f"Completed at {datetime.now().strftime('%H:%M:%S')}")
                                
                                try:
                                    await interaction.message.delete()
                                except:
                                    pass
                                
                                await interaction.followup.send(embed=confirmation_embed, ephemeral=True)
                                
                                channel = self.bot_instance.bot.get_channel(self.bot_instance.submission_channel_id)
                                if channel:
                                    public_embed = discord.Embed(
                                        title="✅ Submission Approved",
                                        description=f"A submission has been approved by {self.user_name}.",
                                        color=discord.Color.green()
                                    )
                                    public_embed.add_field(name="Wallet", value=f"`{self.wallet[:8]}...{self.wallet[-6:]}`", inline=False)
                                    public_embed.set_footer(text=f"Approved at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
                                    await channel.send(embed=public_embed)
                            else:
                                await interaction.followup.send(
                                    f"❌ Failed to approve submission for wallet `{self.wallet}`",
                                    ephemeral=True
                                )
                        
                        @discord.ui.button(label="Reject", style=discord.ButtonStyle.danger, emoji="❌", row=0)
                        async def reject_button(self, interaction: discord.Interaction, button: discord.ui.Button):
                            await interaction.response.defer()
                            
                            success = await self.bot_instance._update_excel_status(self.wallet, "Rejected")
                            
                            if success:
                                # Send confirmation message
                                confirmation_embed = discord.Embed(
                                    title="❌ Submission Rejected",
                                    description=f"**{self.wallet}** has been rejected by {self.user_name}.",
                                    color=discord.Color.red()
                                )
                                confirmation_embed.add_field(name="Status", value="❌ **REJECTED**", inline=True)
                                confirmation_embed.add_field(name="Action", value="❌ Rejected", inline=True)
                                confirmation_embed.set_footer(text=f"Completed at {datetime.now().strftime('%H:%M:%S')}")
                                
                                # Delete the review message
                                try:
                                    await interaction.message.delete()
                                except:
                                    pass
                                
                                # Send confirmation
                                await interaction.followup.send(embed=confirmation_embed, ephemeral=True)
                                
                                # Also send a public notification to the channel
                                channel = self.bot_instance.bot.get_channel(self.bot_instance.submission_channel_id)
                                if channel:
                                    public_embed = discord.Embed(
                                        title="❌ Submission Rejected",
                                        description=f"A submission has been rejected by {self.user_name}.",
                                        color=discord.Color.red()
                                    )
                                    public_embed.add_field(name="Wallet", value=f"`{self.wallet[:8]}...{self.wallet[-6:]}`", inline=False)
                                    public_embed.set_footer(text=f"Rejected at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
                                    await channel.send(embed=public_embed)
                            else:
                                await interaction.followup.send(
                                    f"❌ Failed to reject submission for wallet `{self.wallet}`",
                                    ephemeral=True
                                )
                        
                        @discord.ui.button(label="View Full Details", style=discord.ButtonStyle.secondary, emoji="👁️", row=1)
                        async def view_button(self, interaction: discord.Interaction, button: discord.ui.Button):
                            await interaction.response.defer()
                            
                            record = await self.bot_instance._get_record_details(self.wallet)
                            
                            if not record:
                                await interaction.followup.send(
                                    f"❌ Could not find details for wallet: `{self.wallet}`",
                                    ephemeral=True
                                )
                                return
                            
                            # Create a more detailed embed
                            full_embed = discord.Embed(
                                title=f"📋 Full Submission Details",
                                description=f"**{record.get('Name', 'N/A')}**",
                                color=discord.Color.green()
                            )
                            
                            # Professional Info
                            full_embed.add_field(
                                name="Professional Info",
                                value=f"**Role:** {record.get('Role', 'N/A')}\n"
                                    f"**Injective Role:** {record.get('Injective Role', 'N/A')}\n"
                                    f"**Experience:** {record.get('Experience', 'N/A')}\n"
                                    f"**Education:** {record.get('Education', 'N/A')}\n"
                                    f"**Location:** {record.get('Location', 'N/A')}\n"
                                    f"**Monthly Rate:** {record.get('Monthly Rate', 'N/A')}\n"
                                    f"**Available:** {record.get('Availability', 'N/A')}",
                                inline=False
                            )
                            
                            # Contact info
                            contact_info = [
                                f"**Discord:** {record.get('Discord', '-')}",
                                f"**Email:** {record.get('Email', '-')}",
                                f"**Phone:** {record.get('Phone', '-')}",
                                f"**Telegram:** {record.get('Telegram', '-')}",
                                f"**X/Twitter:** {record.get('X', '-')}",
                                f"**GitHub:** {record.get('Github', '-')}"
                            ]
                            full_embed.add_field(
                                name="Contact Information",
                                value="\n".join(contact_info),
                                inline=False
                            )
                            
                            # Wallet info
                            full_embed.add_field(
                                name="Wallet Information",
                                value=f"**Address:** `{self.wallet}`\n"
                                    f"**Type:** {record.get('Wallet Type', 'N/A')}\n"
                                    f"**NFT Holdings:** {record.get('NFT Holdings', 'N/A')}\n"
                                    f"**Token Holdings:** {record.get('Token Holdings', 'N/A')}",
                                inline=False
                            )
                            
                            # Links
                            links = []
                            portfolio = record.get('Portfolio', '')
                            if portfolio and portfolio != 'N/A' and portfolio != '-':
                                links.append(f"**Portfolio:** [Link]({portfolio})")
                            
                            cv = record.get('CV', '')
                            if cv and cv != 'N/A' and cv != '-':
                                links.append(f"**CV/Resume:** [Download]({cv})")
                            
                            image = record.get('Image url', '')
                            if image and image != 'N/A' and image != '-':
                                links.append(f"**Profile Picture:** [View]({image})")
                            
                            if links:
                                full_embed.add_field(
                                    name="Links",
                                    value="\n".join(links),
                                    inline=False
                                )
                            
                            # Bio
                            bio = record.get('Bio', 'No bio provided')
                            full_embed.add_field(
                                name="Bio",
                                value=f"{bio[:1000]}{'...' if len(bio) > 1000 else ''}",
                                inline=False
                            )
                            
                            full_embed.set_footer(text=f"Wallet: {self.wallet[:8]}...{self.wallet[-6:]}")
                            
                            await interaction.followup.send(embed=full_embed, ephemeral=True)
                        
                        @discord.ui.button(label="Close Review", style=discord.ButtonStyle.secondary, emoji="❌", row=1)
                        async def close_button(self, interaction: discord.Interaction, button: discord.ui.Button):
                            await interaction.response.defer()
                            
                            # Delete the review message
                            try:
                                await interaction.message.delete()
                            except:
                                pass
                            
                            # Send confirmation
                            await interaction.followup.send(
                                "✅ Review closed. Use `/job_open` to review another submission.",
                                ephemeral=True
                            )
                    
                    await interaction.followup.send(
                        embed=detail_embed, 
                        view=ReviewView(self.bot_instance, wallet, interaction.user.name, self.view),
                        ephemeral=True
                    )
            
            class SubmissionView(discord.ui.View):
                def __init__(self, bot_instance, records):
                    super().__init__(timeout=180)
                    self.add_item(SubmissionSelect(bot_instance, records))
                    self.bot_instance = bot_instance
                
                @discord.ui.button(label="Close Menu", style=discord.ButtonStyle.secondary, row=1)
                async def close_button(self, interaction: discord.Interaction, button: discord.ui.Button):
                    await interaction.response.defer()
                    
                    # Delete the menu message
                    try:
                        await interaction.message.delete()
                    except:
                        pass
                    
                    # Send confirmation
                    await interaction.followup.send(
                        "✅ Menu closed. Use `/job_open` to review submissions again.",
                        ephemeral=True
                    )
            
            view = SubmissionView(self, pending_records)
            original_message = await interaction.followup.send(embed=embed, view=view)
            
        except Exception as e:
            print(f"❌ Error in job_open command: {e}")
            traceback.print_exc()
            await interaction.followup.send(
                "An error occurred while opening pending submissions.",
                ephemeral=True
            )
    
    def _ensure_excel_file(self):
        """Ensure the Excel file exists with correct headers, if not then create it."""
        if not os.path.exists(self.excel_file):
            print(f"📄 Creating new Excel file: {self.excel_file}")
            wb = Workbook()
            ws = wb.active
            ws.title = "Submissions"
            
            headers = [
                "Name", "Role", "Injective Role", "Experience", "Education", "Location",
                "Availability", "Monthly Rate", "Skills", "Languages", "Discord", "Email",
                "Phone", "Telegram", "X", "Github", "Wallet Address", "Wallet Type",
                "NFT Holdings", "Token Holdings", "Portfolio", "CV", "Image url", "Bio",
                "Submission date", "Status"
            ]
            
            ws.append(headers)
            wb.save(self.excel_file)
            print(f"✅ Initialized Excel file with headers")
        else:
            try:
                wb = load_workbook(self.excel_file)
                ws = wb.active
                expected_headers = [
                    "Name", "Role", "Injective Role", "Experience", "Education", 
                    "Location", "Availability", "Monthly Rate", "Skills", "Languages",
                    "Discord", "Email", "Phone", "Telegram", "X", "Github",
                    "Wallet Address", "Wallet Type", "NFT Holdings", "Token Holdings",
                    "Portfolio", "CV", "Image url", "Bio", "Submission date", "Status"
                ]
                actual_headers = [cell.value for cell in ws[1]]
                if actual_headers != expected_headers:
                    print(f"📄 Excel headers mismatch. Recreating file...")
                    os.remove(self.excel_file)
                    self._init_excel_file()
                wb.close()
            except Exception as e:
                print(f"❌ Error verifying Excel file: {e}")
                if os.path.exists(self.excel_file):
                    os.remove(self.excel_file)
                self._init_excel_file()

    def _init_excel_file(self):
        """Initialize Excel file with headers"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Submissions"
        
        headers = [
            "Name", "Role", "Injective Role", "Experience", "Education", "Location",
            "Availability", "Monthly Rate", "Skills", "Languages", "Discord", "Email",
            "Phone", "Telegram", "X", "Github", "Wallet Address", "Wallet Type",
            "NFT Holdings", "Token Holdings", "Portfolio", "CV", "Image url", "Bio",
            "Submission date", "Status"
        ]
        
        ws.append(headers)
        wb.save(self.excel_file)
        print(f"✅ Initialized Excel file with headers")

    async def _find_submission_row(self, wallet_address: str) -> Optional[int]:
        """Find the row number of a submission by wallet address."""
        try:
            wb = load_workbook(self.excel_file, read_only=True)
            ws = wb.active
            
            for row in range(2, ws.max_row + 1):
                cell_value = ws[f'Q{row}'].value
                if cell_value and str(cell_value).strip().lower() == wallet_address.strip().lower():
                    wb.close()
                    return row
            
            wb.close()
            return None
            
        except Exception as e:
            print(f"❌ Error searching Excel file: {e}")
            return None
        
    async def _get_all_records(self) -> List[Tuple[str, str]]:
        """Get all records as a list of (Name, Wallet Address) tuples."""
        try:
            wb = load_workbook(self.excel_file, read_only=True)
            ws = wb.active
            
            records = []
            for row in range(2, ws.max_row + 1):
                name = ws[f'A{row}'].value
                wallet = ws[f'Q{row}'].value
                if name and wallet:
                    records.append((name, wallet))
            wb.close()
            return records
        except Exception as e:
            print(f"❌ Error getting all records: {e}")
            return []
    
    async def _get_pending_records(self) -> List[Tuple[str, str]]:
        """Get only pending records as a list of (Name, Wallet Address) tuples."""
        try:
            wb = load_workbook(self.excel_file, read_only=True)
            ws = wb.active
            
            records = []
            for row in range(2, ws.max_row + 1):
                name = ws[f'A{row}'].value
                wallet = ws[f'Q{row}'].value
                status = ws[f'Z{row}'].value
                
                if name and wallet and status == "Pending":
                    records.append((name, wallet))
            wb.close()
            return records
        except Exception as e:
            print(f"❌ Error getting pending records: {e}")
            return []
    
    async def _get_approved_records(self) -> List[Tuple[str, str]]:
        """Get only approved records as a list of (Name, Wallet Address) tuples."""
        try:
            wb = load_workbook(self.excel_file, read_only=True)
            ws = wb.active
            
            records = []
            for row in range(2, ws.max_row + 1):
                name = ws[f'A{row}'].value
                wallet = ws[f'Q{row}'].value
                status = ws[f'Z{row}'].value
                
                if name and wallet and status == "Approved":
                    records.append((name, wallet))
            wb.close()
            return records
        except Exception as e:
            print(f"❌ Error getting approved records: {e}")
            return []
    
    async def _get_all_records_with_details(self) -> List[dict]:
        """Get all records with full details as a list of dictionaries."""
        try:
            wb = load_workbook(self.excel_file, read_only=True)
            ws = wb.active
            
            records = []
            for row in range(2, ws.max_row + 1):
                name = ws[f'A{row}'].value
                wallet = ws[f'Q{row}'].value
                
                if name and wallet:
                    record = {
                        'Name': name,
                        'Role': ws[f'B{row}'].value or 'N/A',
                        'Injective Role': ws[f'C{row}'].value or 'N/A',
                        'Experience': ws[f'D{row}'].value or 'N/A',
                        'Education': ws[f'E{row}'].value or 'N/A',
                        'Location': ws[f'F{row}'].value or 'N/A',
                        'Availability': ws[f'G{row}'].value or 'N/A',
                        'Monthly Rate': ws[f'H{row}'].value or 'N/A',
                        'Skills': ws[f'I{row}'].value or 'N/A',
                        'Languages': ws[f'J{row}'].value or 'N/A',
                        'Discord': ws[f'K{row}'].value or 'N/A',
                        'Email': ws[f'L{row}'].value or 'N/A',
                        'Phone': ws[f'M{row}'].value or 'N/A',
                        'Telegram': ws[f'N{row}'].value or 'N/A',
                        'X': ws[f'O{row}'].value or 'N/A',
                        'Github': ws[f'P{row}'].value or 'N/A',
                        'Wallet Address': wallet,
                        'Wallet Type': ws[f'R{row}'].value or 'N/A',
                        'NFT Holdings': ws[f'S{row}'].value or 'N/A',
                        'Token Holdings': ws[f'T{row}'].value or 'N/A',
                        'Portfolio': ws[f'U{row}'].value or 'N/A',
                        'CV': ws[f'V{row}'].value or 'N/A',
                        'Image url': ws[f'W{row}'].value or 'N/A',
                        'Bio': ws[f'X{row}'].value or 'N/A',
                        'Status': ws[f'Z{row}'].value or 'Pending'
                    }
                    records.append(record)
            wb.close()
            return records
        except Exception as e:
            print(f"❌ Error getting all records with details: {e}")
            return []
    
    async def _get_status_counts(self) -> dict:
        """Get counts of submissions by status."""
        try:
            wb = load_workbook(self.excel_file, read_only=True)
            ws = wb.active
            
            counts = {
                'Approved': 0,
                'Pending': 0,
                'Rejected': 0,
                'Total': 0
            }
            
            for row in range(2, ws.max_row + 1):
                status = ws[f'Z{row}'].value
                if status in counts:
                    counts[status] += 1
                counts['Total'] += 1
            
            wb.close()
            return counts
        except Exception as e:
            print(f"❌ Error getting status counts: {e}")
            return {'Approved': 0, 'Pending': 0, 'Rejected': 0, 'Total': 0}
    
    async def _delete_record(self, wallet_address: str) -> bool:
        """Delete a record from the Excel file by wallet address."""
        try:
            wb = load_workbook(self.excel_file)
            ws = wb.active
            
            row = await self._find_submission_row(wallet_address)
            if not row:
                wb.close()
                return False
            
            ws.delete_rows(row)
            wb.save(self.excel_file)
            wb.close()
            return True
            
        except Exception as e:
            print(f"❌ Error deleting record: {e}")
            return False
    
    async def _update_single_field(self, wallet_address: str, column_name: str, new_value: str) -> bool:
        """Update a single field in the Excel file for a specific wallet"""
        try:
            column_map = {
                'Name': 'A',
                'Role': 'B',
                'Injective Role': 'C',
                'Experience': 'D',
                'Education': 'E',
                'Location': 'F',
                'Availability': 'G',
                'Monthly Rate': 'H',
                'Skills': 'I',
                'Languages': 'J',
                'Discord': 'K',
                'Email': 'L',
                'Phone': 'M',
                'Telegram': 'N',
                'X': 'O',
                'Github': 'P',
                'Wallet Type': 'R',
                'NFT Holdings': 'S',
                'Token Holdings': 'T',
                'Portfolio': 'U',
                'CV': 'V',
                'Image url': 'W',
                'Bio': 'X',
                'Status': 'Z'
            }
            
            if column_name not in column_map:
                return False
            
            wb = load_workbook(self.excel_file)
            ws = wb.active
            
            row = await self._find_submission_row(wallet_address)
            if not row:
                wb.close()
                return False
            
            col = column_map[column_name]
            ws[f'{col}{row}'] = new_value
            
            ws[f'Y{row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            wb.save(self.excel_file)
            wb.close()
            return True
            
        except Exception as e:
            print(f"❌ Error updating field: {e}")
            return False
    
    async def _get_record_details(self, wallet_address: str) -> Optional[dict]:
        """Get full details of a record by wallet address."""
        try:
            wb = load_workbook(self.excel_file, read_only=True)
            ws = wb.active
            
            row = await self._find_submission_row(wallet_address)
            if not row:
                wb.close()
                return None
            
            record = {}
            columns = [
                ('Name', 'A'), ('Role', 'B'), ('Injective Role', 'C'),
                ('Experience', 'D'), ('Education', 'E'), ('Location', 'F'),
                ('Availability', 'G'), ('Monthly Rate', 'H'), ('Skills', 'I'),
                ('Languages', 'J'), ('Discord', 'K'), ('Email', 'L'),
                ('Phone', 'M'), ('Telegram', 'N'), ('X', 'O'),
                ('Github', 'P'), ('Wallet Address', 'Q'), ('Wallet Type', 'R'),
                ('NFT Holdings', 'S'), ('Token Holdings', 'T'), ('Portfolio', 'U'),
                ('CV', 'V'), ('Image url', 'W'), ('Bio', 'X'), ('Status', 'Z')
            ]
            
            for col_name, col_letter in columns:
                value = ws[f'{col_letter}{row}'].value
                record[col_name] = value if value is not None else "N/A"
            
            wb.close()
            return record
            
        except Exception as e:
            print(f"❌ Error getting record details: {e}")
            return None
    
    async def _update_excel_status(self, wallet_address: str, new_status: str) -> bool:
        """Update the status column in Excel for a specific wallet"""
        try:
            row = await self._find_submission_row(wallet_address)
            if not row:
                print(f"[EXCEL ERROR] No row found for wallet: {wallet_address}")
                return False
            
            wb = load_workbook(self.excel_file)
            ws = wb.active
            
            # Update the status
            ws[f'Z{row}'] = new_status
            
            # Save the file
            wb.save(self.excel_file)
            wb.close()
            
            print(f"[EXCEL SUCCESS] Updated {wallet_address} status to '{new_status}'")
            return True
            
        except Exception as e:
            print(f"[EXCEL ERROR] Error updating Excel status: {e}")
            traceback.print_exc()
            return False
    
    async def job_open_command(self, interaction: discord.Interaction):
        """
        Open pending submissions to review.
        Shows dropdown with all pending submissions (name and role).
        When selected, shows embed with details and Approve/Reject/Close buttons.
        Updates Excel status accordingly.
        """
        await interaction.response.defer()
        
        try:
            # Get pending submissions from Excel
            pending_records = await self._get_pending_records()
            
            if not pending_records:
                embed = discord.Embed(
                    title="📭 Pending Submissions",
                    description="No submissions waiting for review.",
                    color=discord.Color.green()
                )
                await interaction.followup.send(embed=embed)
                return
            
            # Create a selection menu
            embed = discord.Embed(
                title="📬 Pending Submissions",
                description=f"Select a submission to review:",
                color=discord.Color.blue()
            )
            embed.add_field(name="📋 Total Pending", value=f"{len(pending_records)} submission(s)", inline=False)
            embed.set_footer(text="Select a submission from the dropdown below")
            
            # Create dropdown with pending submissions
            class SubmissionSelect(discord.ui.Select):
                def __init__(self, bot_instance, records):
                    options = []
                    for name, wallet in records[:25]:  # Discord limit: 25 options
                        # Get role for the description - simplified to avoid blocking
                        options.append(discord.SelectOption(
                            label=f"{name[:90]}",
                            description=f"Click to review",
                            value=wallet,
                            emoji="📄"
                        ))
                    
                    super().__init__(
                        placeholder="Choose a submission to review...",
                        min_values=1,
                        max_values=1,
                        options=options
                    )
                    self.bot_instance = bot_instance
                    self.records = records  # Store records for later use
                
                async def callback(self, interaction: discord.Interaction):
                    await interaction.response.defer()
                    wallet = self.values[0]
                    
                    # Get record details
                    record = await self.bot_instance._get_record_details(wallet)
                    
                    if not record:
                        await interaction.followup.send(
                            f"❌ Could not find details for wallet: `{wallet}`",
                            ephemeral=True
                        )
                        return
                    
                    # Create embed with submission details
                    detail_embed = discord.Embed(
                        title=f"📄 Submission Review",
                        description=f"**{record.get('Name', 'N/A')}** - {record.get('Role', 'N/A')}",
                        color=discord.Color.blue()
                    )
                    
                    # Basic info
                    detail_embed.add_field(
                        name="Basic Info",
                        value=f"**Experience:** {record.get('Experience', 'N/A')}\n"
                            f"**Education:** {record.get('Education', 'N/A')}\n"
                            f"**Location:** {record.get('Location', 'N/A')}\n"
                            f"**Available:** {record.get('Availability', 'N/A')}",
                        inline=False
                    )
                    
                    # Wallet info
                    detail_embed.add_field(
                        name="Wallet",
                        value=f"`{wallet}`",
                        inline=False
                    )
                    
                    # Skills
                    skills = record.get('Skills', 'N/A')
                    detail_embed.add_field(
                        name="Skills",
                        value=skills[:500] + "..." if len(skills) > 500 else skills,
                        inline=False
                    )
                    
                    # Languages
                    languages = record.get('Languages', 'N/A')
                    detail_embed.add_field(
                        name="Languages",
                        value=languages[:200] + "..." if len(languages) > 200 else languages,
                        inline=False
                    )
                    
                    detail_embed.set_footer(text=f"Wallet: {wallet[:8]}...{wallet[-6:]}")
                    
                    # Create review buttons
                    class ReviewView(discord.ui.View):
                        def __init__(self, bot_instance, wallet, user_name):
                            super().__init__(timeout=180)
                            self.bot_instance = bot_instance
                            self.wallet = wallet
                            self.user_name = user_name
                        
                        @discord.ui.button(label="Approve", style=discord.ButtonStyle.success, emoji="✅", row=0)
                        async def approve_button(self, interaction: discord.Interaction, button: discord.ui.Button):
                            await interaction.response.defer()
                            
                            success = await self.bot_instance._update_excel_status(self.wallet, "Approved")
                            
                            if success:
                                await interaction.followup.send(
                                    f"✅ Submission approved for wallet `{self.wallet}`!",
                                    ephemeral=True
                                )
                                
                                # Update original message
                                detail_embed.title = f"✅ Submission Approved"
                                detail_embed.color = discord.Color.green()
                                detail_embed.set_footer(text=f"Approved by {self.user_name}")
                                
                                # Create new view with option to review another
                                class AnotherView(discord.ui.View):
                                    def __init__(self, bot_instance):
                                        super().__init__(timeout=30)
                                        self.bot_instance = bot_instance
                                    
                                    @discord.ui.button(label="Review Another", style=discord.ButtonStyle.primary)
                                    async def another_button(self, interaction: discord.Interaction, button: discord.ui.Button):
                                        await interaction.response.defer()
                                        # Re-run the job_open command
                                        await self.bot_instance.job_open_command(interaction)
                                    
                                    @discord.ui.button(label="Close", style=discord.ButtonStyle.secondary)
                                    async def close_button(self, interaction: discord.Interaction, button: discord.ui.Button):
                                        await interaction.response.defer()
                                        await interaction.message.delete()
                                        await interaction.followup.send("✅ Review closed.", ephemeral=True)
                                
                                await interaction.message.edit(embed=detail_embed, view=AnotherView(self.bot_instance))
                            else:
                                await interaction.followup.send(
                                    f"❌ Failed to approve submission for wallet `{self.wallet}`",
                                    ephemeral=True
                                )
                        
                        @discord.ui.button(label="Reject", style=discord.ButtonStyle.danger, emoji="❌", row=0)
                        async def reject_button(self, interaction: discord.Interaction, button: discord.ui.Button):
                            await interaction.response.defer()
                            
                            success = await self.bot_instance._update_excel_status(self.wallet, "Rejected")
                            
                            if success:
                                await interaction.followup.send(
                                    f"❌ Submission rejected for wallet `{self.wallet}`!",
                                    ephemeral=True
                                )
                                
                                # Update original message
                                detail_embed.title = f"❌ Submission Rejected"
                                detail_embed.color = discord.Color.red()
                                detail_embed.set_footer(text=f"Rejected by {self.user_name}")
                                
                                # Create new view with option to review another
                                class AnotherView(discord.ui.View):
                                    def __init__(self, bot_instance):
                                        super().__init__(timeout=30)
                                        self.bot_instance = bot_instance
                                    
                                    @discord.ui.button(label="Review Another", style=discord.ButtonStyle.primary)
                                    async def another_button(self, interaction: discord.Interaction, button: discord.ui.Button):
                                        await interaction.response.defer()
                                        # Re-run the job_open command
                                        await self.bot_instance.job_open_command(interaction)
                                    
                                    @discord.ui.button(label="Close", style=discord.ButtonStyle.secondary)
                                    async def close_button(self, interaction: discord.Interaction, button: discord.ui.Button):
                                        await interaction.response.defer()
                                        await interaction.message.delete()
                                        await interaction.followup.send("✅ Review closed.", ephemeral=True)
                                
                                await interaction.message.edit(embed=detail_embed, view=AnotherView(self.bot_instance))
                            else:
                                await interaction.followup.send(
                                    f"❌ Failed to reject submission for wallet `{self.wallet}`",
                                    ephemeral=True
                                )
                        
                        @discord.ui.button(label="View Full Details", style=discord.ButtonStyle.secondary, emoji="👁️", row=1)
                        async def view_button(self, interaction: discord.Interaction, button: discord.ui.Button):
                            await interaction.response.defer()
                            
                            record = await self.bot_instance._get_record_details(self.wallet)
                            
                            if not record:
                                await interaction.followup.send(
                                    f"❌ Could not find details for wallet: `{self.wallet}`",
                                    ephemeral=True
                                )
                                return
                            
                            # Create a more detailed embed
                            full_embed = discord.Embed(
                                title=f"📋 Full Submission Details",
                                description=f"**{record.get('Name', 'N/A')}**",
                                color=discord.Color.green()
                            )
                            
                            # Professional Info
                            full_embed.add_field(
                                name="Professional Info",
                                value=f"**Role:** {record.get('Role', 'N/A')}\n"
                                    f"**Injective Role:** {record.get('Injective Role', 'N/A')}\n"
                                    f"**Experience:** {record.get('Experience', 'N/A')}\n"
                                    f"**Education:** {record.get('Education', 'N/A')}\n"
                                    f"**Location:** {record.get('Location', 'N/A')}\n"
                                    f"**Monthly Rate:** {record.get('Monthly Rate', 'N/A')}\n"
                                    f"**Available:** {record.get('Availability', 'N/A')}",
                                inline=False
                            )
                            
                            # Contact info
                            contact_info = [
                                f"**Discord:** {record.get('Discord', '-')}",
                                f"**Email:** {record.get('Email', '-')}",
                                f"**Phone:** {record.get('Phone', '-')}",
                                f"**Telegram:** {record.get('Telegram', '-')}",
                                f"**X/Twitter:** {record.get('X', '-')}",
                                f"**GitHub:** {record.get('Github', '-')}"
                            ]
                            full_embed.add_field(
                                name="Contact Information",
                                value="\n".join(contact_info),
                                inline=False
                            )
                            
                            # Wallet info
                            full_embed.add_field(
                                name="Wallet Information",
                                value=f"**Address:** `{self.wallet}`\n"
                                    f"**Type:** {record.get('Wallet Type', 'N/A')}\n"
                                    f"**NFT Holdings:** {record.get('NFT Holdings', 'N/A')}\n"
                                    f"**Token Holdings:** {record.get('Token Holdings', 'N/A')}",
                                inline=False
                            )
                            
                            # Links
                            links = []
                            portfolio = record.get('Portfolio', '')
                            if portfolio and portfolio != 'N/A' and portfolio != '-':
                                links.append(f"**Portfolio:** [Link]({portfolio})")
                            
                            cv = record.get('CV', '')
                            if cv and cv != 'N/A' and cv != '-':
                                links.append(f"**CV/Resume:** [Download]({cv})")
                            
                            image = record.get('Image url', '')
                            if image and image != 'N/A' and image != '-':
                                links.append(f"**Profile Picture:** [View]({image})")
                            
                            if links:
                                full_embed.add_field(
                                    name="Links",
                                    value="\n".join(links),
                                    inline=False
                                )
                            
                            # Bio
                            bio = record.get('Bio', 'No bio provided')
                            full_embed.add_field(
                                name="Bio",
                                value=f"{bio[:1000]}{'...' if len(bio) > 1000 else ''}",
                                inline=False
                            )
                            
                            full_embed.set_footer(text=f"Wallet: {self.wallet[:8]}...{self.wallet[-6:]}")
                            
                            await interaction.followup.send(embed=full_embed, ephemeral=True)
                        
                        @discord.ui.button(label="Close", style=discord.ButtonStyle.danger, emoji="❌", row=1)
                        async def close_button(self, interaction: discord.Interaction, button: discord.ui.Button):
                            await interaction.response.defer()
                            await interaction.message.delete()
                            await interaction.followup.send("✅ Review closed.", ephemeral=True)
                    
                    await interaction.followup.send(
                        embed=detail_embed, 
                        view=ReviewView(self.bot_instance, wallet, interaction.user.name),
                        ephemeral=True
                    )
            
            class SubmissionView(discord.ui.View):
                def __init__(self, bot_instance, records):
                    super().__init__(timeout=180)
                    self.add_item(SubmissionSelect(bot_instance, records))
                
                @discord.ui.button(label="Close Menu", style=discord.ButtonStyle.secondary, row=1)
                async def close_button(self, interaction: discord.Interaction, button: discord.ui.Button):
                    await interaction.response.defer()
                    await interaction.message.delete()
                    await interaction.followup.send("✅ Menu closed.", ephemeral=True)
            
            view = SubmissionView(self, pending_records)
            await interaction.followup.send(embed=embed, view=view)
            
        except Exception as e:
            print(f"❌ Error in job_open command: {e}")
            traceback.print_exc()
            await interaction.followup.send(
                "An error occurred while opening pending submissions.",
                ephemeral=True
            )
            class SubmissionView(discord.ui.View):
                def __init__(self, bot_instance, records):
                    super().__init__(timeout=180)
                    self.add_item(SubmissionSelect(bot_instance, records))
                
                @discord.ui.button(label="Close Menu", style=discord.ButtonStyle.secondary, row=1)
                async def close_button(self, interaction: discord.Interaction, button: discord.ui.Button):
                    await interaction.response.defer()
                    await interaction.message.delete()
                    await interaction.followup.send("✅ Menu closed.", ephemeral=True)
            
            view = SubmissionView(self, pending_records)
            await interaction.followup.send(embed=embed, view=view)
            
        except Exception as e:
            print(f"❌ Error in job_open command: {e}")
            traceback.print_exc()
            await interaction.followup.send(
                "An error occurred while opening pending submissions.",
                ephemeral=True
            )
    
    async def job_open_prefix_command(self, ctx: commands.Context):
        """Open pending submissions for review (prefix command)"""
        try:
            # Get pending submissions from Excel
            pending_records = await self._get_pending_records()
            
            if not pending_records:
                embed = discord.Embed(
                    title="📭 Pending Submissions",
                    description="No submissions waiting for review.",
                    color=discord.Color.green()
                )
                await ctx.send(embed=embed)
                return
            
            embed = discord.Embed(
                title="📬 Pending Submissions",
                description=f"**{len(pending_records)}** submission(s) waiting for review",
                color=discord.Color.blue()
            )
            
            # List pending submissions
            submissions_list = []
            for i, (name, wallet) in enumerate(pending_records[:10], 1):
                submissions_list.append(f"**{i}. {name}**\n   `{wallet}`")
            
            if submissions_list:
                embed.add_field(
                    name="Pending Submissions",
                    value="\n\n".join(submissions_list),
                    inline=False
                )
            
            embed.set_footer(text="Use the slash command /job_open for interactive review")
            
            await ctx.send(embed=embed)
            
        except Exception as e:
            print(f"❌ Error in job_open prefix command: {e}")
            await ctx.send("An error occurred while fetching pending submissions.")
    
    async def job_status_command(self, interaction: discord.Interaction):
        """Show current submission status - how many approved, rejected, and pending"""
        await interaction.response.defer(ephemeral=True)
        
        try:
            counts = await self._get_status_counts()
            
            embed = discord.Embed(
                title="📊 Submission Status",
                color=discord.Color.blue()
            )
            
            embed.add_field(
                name="✅ Approved",
                value=f"**{counts['Approved']}** submission(s)",
                inline=True
            )
            
            embed.add_field(
                name="⏳ Pending",
                value=f"**{counts['Pending']}** submission(s)",
                inline=True
            )
            
            embed.add_field(
                name="❌ Rejected",
                value=f"**{counts['Rejected']}** submission(s)",
                inline=True
            )
            
            embed.add_field(
                name="📊 Total",
                value=f"**{counts['Total']}** submission(s)",
                inline=False
            )
            
            embed.set_footer(text="Use /job_open to review pending submissions")
            
            await interaction.followup.send(embed=embed, ephemeral=True)
            
        except Exception as e:
            print(f"❌ Error in job_status command: {e}")
            await interaction.followup.send(
                "An error occurred while fetching status.",
                ephemeral=True
            )
    
    async def job_status_prefix_command(self, ctx: commands.Context):
        """Show current submission status (prefix command)"""
        try:
            counts = await self._get_status_counts()
            
            embed = discord.Embed(
                title="📊 Submission Status",
                color=discord.Color.blue()
            )
            
            embed.add_field(
                name="✅ Approved",
                value=f"**{counts['Approved']}** submission(s)",
                inline=True
            )
            
            embed.add_field(
                name="⏳ Pending",
                value=f"**{counts['Pending']}** submission(s)",
                inline=True
            )
            
            embed.add_field(
                name="❌ Rejected",
                value=f"**{counts['Rejected']}** submission(s)",
                inline=True
            )
            
            embed.add_field(
                name="📊 Total",
                value=f"**{counts['Total']}** submission(s)",
                inline=False
            )
            
            embed.set_footer(text="Use /job_open to review pending submissions")
            
            await ctx.send(embed=embed)
            
        except Exception as e:
            print(f"❌ Error in job_status prefix command: {e}")
            await ctx.send("An error occurred while fetching status.")
    
    async def show_command(self, interaction: discord.Interaction):
        """Show all approved submissions with pagination"""
        await interaction.response.defer()
        
        try:
            records = await self._get_approved_records()
            
            if not records:
                await interaction.followup.send("No approved records found in the database.")
                return
            
            class PaginatedView(View):
                def __init__(self, records: List[Tuple[str, str]], page: int = 0):
                    super().__init__(timeout=180)
                    self.records = records
                    self.page = page
                    self.records_per_page = 5
                    self.total_pages = (len(records) + self.records_per_page - 1) // self.records_per_page
                
                def create_embed(self):
                    """Create embed for current page"""
                    start_idx = self.page * self.records_per_page
                    end_idx = min(start_idx + self.records_per_page, len(self.records))
                    
                    embed = discord.Embed(
                        title="✅ Approved Submissions",
                        description=f"Showing approved submissions {start_idx + 1}-{end_idx} of {len(self.records)}",
                        color=discord.Color.green()
                    )
                    
                    for i in range(start_idx, end_idx):
                        name, wallet = self.records[i]
                        embed.add_field(
                            name=f"{name}",
                            value=f"`{wallet}`",
                            inline=False
                        )
                    
                    embed.set_footer(text=f"Page {self.page + 1}/{self.total_pages} | Use buttons to navigate")
                    return embed
                
                async def update_buttons(self):
                    """Update button states"""
                    for child in self.children:
                        if child.custom_id == "prev":
                            child.disabled = self.page == 0
                        elif child.custom_id == "next":
                            child.disabled = self.page == self.total_pages - 1
                
                @discord.ui.button(label="◀ Previous", style=discord.ButtonStyle.primary, custom_id="prev", disabled=True)
                async def previous_button(self, interaction: discord.Interaction, button: Button):
                    await interaction.response.defer()
                    if self.page > 0:
                        self.page -= 1
                        embed = self.create_embed()
                        await self.update_buttons()
                        await interaction.edit_original_response(embed=embed, view=self)
                
                @discord.ui.button(label="Next ▶", style=discord.ButtonStyle.primary, custom_id="next")
                async def next_button(self, interaction: discord.Interaction, button: Button):
                    await interaction.response.defer()
                    if self.page < self.total_pages - 1:
                        self.page += 1
                        embed = self.create_embed()
                        await self.update_buttons()
                        await interaction.edit_original_response(embed=embed, view=self)
                
                @discord.ui.button(label="❌ Close", style=discord.ButtonStyle.danger)
                async def close_button(self, interaction: discord.Interaction, button: Button):
                    await interaction.response.defer()
                    await interaction.edit_original_response(content="View closed.", embed=None, view=None)
                    self.stop()
            
            view = PaginatedView(records)
            
            embed = view.create_embed()
            await view.update_buttons()
            
            await interaction.followup.send(embed=embed, view=view)
            
        except Exception as e:
            print(f"❌ Error in show command: {e}")
            traceback.print_exc()
            await interaction.followup.send(
                "An error occurred while fetching approved records.",
                ephemeral=True
            )
    
    async def show_prefix_command(self, ctx: commands.Context):
        """Show all approved submissions (prefix command)"""
        try:
            records = await self._get_approved_records()
            
            if not records:
                await ctx.send("No approved records found in the database.")
                return
            
            class PaginatedView(View):
                def __init__(self, records: List[Tuple[str, str]], page: int = 0):
                    super().__init__(timeout=180)
                    self.records = records
                    self.page = page
                    self.records_per_page = 5
                    self.total_pages = (len(records) + self.records_per_page - 1) // self.records_per_page
                
                def create_embed(self):
                    """Create embed for current page"""
                    start_idx = self.page * self.records_per_page
                    end_idx = min(start_idx + self.records_per_page, len(self.records))
                    
                    embed = discord.Embed(
                        title="✅ Approved Submissions",
                        description=f"Showing approved submissions {start_idx + 1}-{end_idx} of {len(self.records)}",
                        color=discord.Color.green()
                    )
                    
                    for i in range(start_idx, end_idx):
                        name, wallet = self.records[i]
                        embed.add_field(
                            name=f"{name}",
                            value=f"`{wallet}`",
                            inline=False
                        )
                    
                    embed.set_footer(text=f"Page {self.page + 1}/{self.total_pages} | Use buttons to navigate")
                    return embed
                
                async def update_buttons(self):
                    """Update button states"""
                    for child in self.children:
                        if child.custom_id == "prev_prefix":
                            child.disabled = self.page == 0
                        elif child.custom_id == "next_prefix":
                            child.disabled = self.page == self.total_pages - 1
                
                @discord.ui.button(label="◀ Previous", style=discord.ButtonStyle.primary, custom_id="prev_prefix", disabled=True)
                async def previous_button(self, interaction: discord.Interaction, button: Button):
                    await interaction.response.defer()
                    if self.page > 0:
                        self.page -= 1
                        embed = self.create_embed()
                        await self.update_buttons()
                        await interaction.edit_original_response(embed=embed, view=self)
                
                @discord.ui.button(label="Next ▶", style=discord.ButtonStyle.primary, custom_id="next_prefix")
                async def next_button(self, interaction: discord.Interaction, button: Button):
                    await interaction.response.defer()
                    if self.page < self.total_pages - 1:
                        self.page += 1
                        embed = self.create_embed()
                        await self.update_buttons()
                        await interaction.edit_original_response(embed=embed, view=self)
                
                @discord.ui.button(label="❌ Close", style=discord.ButtonStyle.danger)
                async def close_button(self, interaction: discord.Interaction, button: Button):
                    await interaction.response.defer()
                    await interaction.edit_original_response(content="View closed.", embed=None, view=None)
                    self.stop()
            
            view = PaginatedView(records)
            embed = view.create_embed()
            await view.update_buttons()
            
            await ctx.send(embed=embed, view=view)
            
        except Exception as e:
            print(f"❌ Error in show prefix command: {e}")
            traceback.print_exc()
            await ctx.send("An error occurred while fetching approved records.")
    
    async def change_command(self, interaction: discord.Interaction, wallet_address: str, column_name: str, new_value: str):
        """Change a specific field in a submission"""
        await interaction.response.defer()
        
        valid_columns = [
            'Name', 'Role', 'Injective Role', 'Experience', 'Education',
            'Location', 'Availability', 'Monthly Rate', 'Skills', 'Languages',
            'Discord', 'Email', 'Phone', 'Telegram', 'X', 'Github',
            'Wallet Type', 'NFT Holdings', 'Token Holdings', 'Portfolio',
            'CV', 'Image url', 'Bio', 'Status'
        ]
        
        if column_name not in valid_columns:
            await interaction.followup.send(
                f"Invalid column name. Valid columns are:\n{', '.join(valid_columns)}"
            )
            return
        
        row = await self._find_submission_row(wallet_address)
        if not row:
            await interaction.followup.send(
                f"No record found for wallet address: `{wallet_address}`"
            )
            return
        
        record = await self._get_record_details(wallet_address)
        if not record:
            await interaction.followup.send(
                "Could not retrieve record details."
            )
            return
        
        old_value = record.get(column_name, "N/A")
        
        success = await self._update_single_field(wallet_address, column_name, new_value)
        
        if success:
            embed = discord.Embed(
                title="Field Updated Successfully",
                color=discord.Color.green()
            )
            embed.add_field(name="Wallet Address", value=f"`{wallet_address}`", inline=False)
            embed.add_field(name="Field", value=column_name, inline=True)
            embed.add_field(name="Old Value", value=str(old_value)[:100], inline=True)
            embed.add_field(name="New Value", value=str(new_value)[:100], inline=True)
            embed.set_footer(text=f"Updated by {interaction.user.name}")
            
            await interaction.followup.send(embed=embed)
        else:
            await interaction.followup.send(
                "Failed to update the field. Please check the inputs and try again."
            )
    
    async def change_prefix_command(self, ctx: commands.Context, wallet_address: str, column_name: str, new_value: str):
        """Change a specific field for a wallet address (prefix command)"""
        valid_columns = [
            'Name', 'Role', 'Injective Role', 'Experience', 'Education',
            'Location', 'Availability', 'Monthly Rate', 'Skills', 'Languages',
            'Discord', 'Email', 'Phone', 'Telegram', 'X', 'Github',
            'Wallet Type', 'NFT Holdings', 'Token Holdings', 'Portfolio',
            'CV', 'Image url', 'Bio', 'Status'
        ]
        
        if column_name not in valid_columns:
            await ctx.send(
                f"Invalid column name. Valid columns are:\n{', '.join(valid_columns)}"
            )
            return
        
        row = await self._find_submission_row(wallet_address)
        if not row:
            await ctx.send(f"No record found for wallet address: `{wallet_address}`")
            return
        
        record = await self._get_record_details(wallet_address)
        if not record:
            await ctx.send("Could not retrieve record details.")
            return
        
        old_value = record.get(column_name, "N/A")
        
        processing_msg = await ctx.send("Updating field...")
        
        success = await self._update_single_field(wallet_address, column_name, new_value)
        
        if success:
            embed = discord.Embed(
                title="Field Updated Successfully",
                color=discord.Color.green()
            )
            embed.add_field(name="Wallet Address", value=f"`{wallet_address}`", inline=False)
            embed.add_field(name="Field", value=column_name, inline=True)
            embed.add_field(name="Old Value", value=str(old_value)[:100], inline=True)
            embed.add_field(name="New Value", value=str(new_value)[:100], inline=True)
            embed.set_footer(text=f"Updated by {ctx.author.name}")
            
            await processing_msg.delete()
            await ctx.send(embed=embed)
        else:
            await processing_msg.edit(content="Failed to update the field. Please check the inputs and try again.")
    
    async def remove_command(self, interaction: discord.Interaction, wallet_address: str):
        """Remove a submission from the Excel file"""
        await interaction.response.defer()
        
        row = await self._find_submission_row(wallet_address)
        if not row:
            await interaction.followup.send(
                f"No record found for wallet address: `{wallet_address}`"
            )
            return
        
        record = await self._get_record_details(wallet_address)
        if not record:
            await interaction.followup.send(
                "Could not retrieve record details."
            )
            return
        
        embed = discord.Embed(
            title="Confirm Deletion",
            description=f"You are about to delete the record for:\n**{record.get('Name', 'Unknown')}**",
            color=discord.Color.red()
        )
        embed.add_field(name="Wallet Address", value=f"`{wallet_address}`", inline=False)
        embed.add_field(name="Role", value=record.get('Role', 'N/A'), inline=True)
        embed.add_field(name="Status", value=record.get('Status', 'N/A'), inline=True)
        embed.set_footer(text="This action cannot be undone!")
        
        class ConfirmView(View):
            def __init__(self, bot_instance, wallet, is_slash=True):
                super().__init__(timeout=60)
                self.bot_instance = bot_instance
                self.wallet = wallet
                self.is_slash = is_slash
                self.confirmed = False
            
            @discord.ui.button(label="Confirm Delete", style=discord.ButtonStyle.danger)
            async def confirm_delete(self, interaction: discord.Interaction, button: Button):
                await interaction.response.defer()
                self.confirmed = True
                
                success = await self.bot_instance._delete_record(self.wallet)
                
                if success:
                    embed = discord.Embed(
                        title="Record Deleted Successfully",
                        description=f"Record for wallet `{self.wallet}` has been deleted.",
                        color=discord.Color.red()
                    )
                    
                    await interaction.edit_original_response(embed=embed, view=None)
                else:
                    await interaction.edit_original_response(
                        content="Failed to delete the record.",
                        embed=None,
                        view=None
                    )
                
                self.stop()
            
            @discord.ui.button(label="Cancel", style=discord.ButtonStyle.secondary)
            async def cancel(self, interaction: discord.Interaction, button: Button):
                await interaction.response.defer()
                await interaction.edit_original_response(
                    content="Deletion cancelled.",
                    embed=None,
                    view=None
                )
                self.stop()
        
        view = ConfirmView(self, wallet_address, is_slash=True)
        await interaction.followup.send(embed=embed, view=view)
    
    async def remove_prefix_command(self, ctx: commands.Context, wallet_address: str):
        """Remove a record by wallet address (prefix command)"""
        row = await self._find_submission_row(wallet_address)
        if not row:
            await ctx.send(f"No record found for wallet address: `{wallet_address}`")
            return
        
        record = await self._get_record_details(wallet_address)
        if not record:
            await ctx.send("Could not retrieve record details.")
            return
        
        embed = discord.Embed(
            title="Confirm Deletion",
            description=f"You are about to delete the record for:\n**{record.get('Name', 'Unknown')}**",
            color=discord.Color.red()
        )
        embed.add_field(name="Wallet Address", value=f"`{wallet_address}`", inline=False)
        embed.add_field(name="Role", value=record.get('Role', 'N/A'), inline=True)
        embed.add_field(name="Status", value=record.get('Status', 'N/A'), inline=True)
        embed.set_footer(text="This action cannot be undone!")
        
        class ConfirmView(View):
            def __init__(self, bot_instance, wallet, is_slash=False):
                super().__init__(timeout=60)
                self.bot_instance = bot_instance
                self.wallet = wallet
                self.is_slash = is_slash
                self.confirmed = False
            
            @discord.ui.button(label="Confirm Delete", style=discord.ButtonStyle.danger)
            async def confirm_delete(self, interaction: discord.Interaction, button: Button):
                await interaction.response.defer()
                self.confirmed = True
                
                success = await self.bot_instance._delete_record(self.wallet)
                
                if success:
                    embed = discord.Embed(
                        title="Record Deleted Successfully",
                        description=f"Record for wallet `{self.wallet}` has been deleted.",
                        color=discord.Color.red()
                    )
                    
                    await interaction.message.edit(embed=embed, view=None)
                    await interaction.response.send_message("Record deleted!", ephemeral=True)
                else:
                    await interaction.message.edit(
                        content="Failed to delete the record.",
                        embed=None,
                        view=None
                    )
                    await interaction.response.send_message("Failed to delete record.", ephemeral=True)
                
                self.stop()
            
            @discord.ui.button(label="Cancel", style=discord.ButtonStyle.secondary)
            async def cancel(self, interaction: discord.Interaction, button: Button):
                await interaction.response.defer()
                await interaction.message.edit(
                    content="Deletion cancelled.",
                    embed=None,
                    view=None
                )
                await interaction.response.send_message("Deletion cancelled.", ephemeral=True)
                self.stop()
        
        view = ConfirmView(self, wallet_address, is_slash=False)
        confirm_msg = await ctx.send(embed=embed, view=view)
    
    async def column_names_command(self, interaction: discord.Interaction):
        """List all available column names for reference"""
        await interaction.response.defer()
        
        try:
            valid_columns = [
                'Name', 'Role', 'Injective Role', 'Experience', 'Education',
                'Location', 'Availability', 'Monthly Rate', 'Skills', 'Languages',
                'Discord', 'Email', 'Phone', 'Telegram', 'X', 'Github',
                'Wallet Type', 'NFT Holdings', 'Token Holdings', 'Portfolio',
                'CV', 'Image url', 'Bio', 'Status'
            ]
            
            embed = discord.Embed(
                title="📋 Available Column Names",
                description="These are the columns you can modify using the /change command:",
                color=discord.Color.blue()
            )
            
            embed.add_field(
                name="Basic Information",
                value="• Name\n• Role\n• Injective Role\n• Experience\n• Education\n• Location\n• Availability\n• Monthly Rate",
                inline=True
            )
            
            embed.add_field(
                name="Skills & Languages",
                value="• Skills\n• Languages",
                inline=True
            )
            
            embed.add_field(
                name="Contact Information",
                value="• Discord\n• Email\n• Phone\n• Telegram\n• X\n• Github",
                inline=True
            )
            
            embed.add_field(
                name="Blockchain Details",
                value="• Wallet Type\n• NFT Holdings\n• Token Holdings",
                inline=True
            )
            
            embed.add_field(
                name="Portfolio & Media",
                value="• Portfolio\n• CV\n• Image url\n• Bio",
                inline=True
            )
            
            embed.add_field(
                name="Status",
                value="• Status",
                inline=True
            )
            
            embed.set_footer(text="Use /change <wallet> <column> <value> to modify these fields")
            
            await interaction.followup.send(embed=embed)
            
        except Exception as e:
            print(f"❌ Error in column_names command: {e}")
            traceback.print_exc()
            await interaction.followup.send(
                "An error occurred while fetching column names."
            )
    
    async def column_names_prefix_command(self, ctx: commands.Context):
        """List all available column names for reference (prefix command)"""
        try:
            valid_columns = [
                'Name', 'Role', 'Injective Role', 'Experience', 'Education',
                'Location', 'Availability', 'Monthly Rate', 'Skills', 'Languages',
                'Discord', 'Email', 'Phone', 'Telegram', 'X', 'Github',
                'Wallet Type', 'NFT Holdings', 'Token Holdings', 'Portfolio',
                'CV', 'Image url', 'Bio', 'Status'
            ]
            
            embed = discord.Embed(
                title="📋 Available Column Names",
                description="These are the columns you can modify using the !change command:",
                color=discord.Color.blue()
            )
            
            embed.add_field(
                name="Basic Information",
                value="• Name\n• Role\n• Injective Role\n• Experience\n• Education\n• Location\n• Availability\n• Monthly Rate",
                inline=True
            )
            
            embed.add_field(
                name="Skills & Languages",
                value="• Skills\n• Languages",
                inline=True
            )
            
            embed.add_field(
                name="Contact Information",
                value="• Discord\n• Email\n• Phone\n• Telegram\n• X\n• Github",
                inline=True
            )
            
            embed.add_field(
                name="Blockchain Details",
                value="• Wallet Type\n• NFT Holdings\n• Token Holdings",
                inline=True
            )
            
            embed.add_field(
                name="Portfolio & Media",
                value="• Portfolio\n• CV\n• Image url\n• Bio",
                inline=True
            )
            
            embed.add_field(
                name="Status",
                value="• Status",
                inline=True
            )
            
            embed.set_footer(text="Use !change <wallet> <column> <value> to modify these fields")
            
            await ctx.send(embed=embed)
            
        except Exception as e:
            print(f"❌ Error in column_names prefix command: {e}")
            await ctx.send("An error occurred while fetching column names.")
    
    async def _save_new_submission(self, data: dict) -> bool:
        """Save a new submission to Excel with status "Pending" (not "Queued")"""
        try:
            wallet = data.get('walletAddress', '').strip()
            
            print(f"[EXCEL SAVE] Saving submission for wallet: {wallet}")
            
            existing_row = await self._find_submission_row(wallet)
            if existing_row:
                print(f"⚠️ [EXCEL SAVE] Wallet {wallet} already exists in row {existing_row}. Updating instead.")
                return await self._update_existing_submission(data, existing_row)
            
            row = [
                data.get('name', '').strip(),
                data.get('role', '').strip(),
                data.get('injectiveRole', '').strip(),
                data.get('experience', '').strip(),
                data.get('education', '').strip(),
                data.get('location', '').strip(),
                'Yes' if data.get('available', False) else 'No',
                data.get('monthlyRate', '').strip(),
                ', '.join(data.get('skills', [])),
                ', '.join(data.get('languages', [])),
                data.get('discord', '').strip(),
                data.get('email', '').strip(),
                data.get('phone', '').strip() or '-',
                data.get('telegram', '').strip() or '-',
                data.get('X', '').strip() or '-',
                data.get('github', '').strip() or '-',
                wallet,
                data.get('walletType', '').strip(),
                data.get('nftHold', '').strip(),
                data.get('tokenHold', '').strip(),
                data.get('portfolio', '').strip() or '-',
                data.get('cv', '').strip(),
                data.get('profilePicture', '').strip(),
                data.get('bio', '').strip(),
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                "Pending"
            ]
            
            wb = load_workbook(self.excel_file)
            ws = wb.active
            ws.append(row)
            wb.save(self.excel_file)
            wb.close()
            print(f"✅ Saved new submission to Excel for wallet: {wallet} (Status: Pending)")
            return True
            
        except Exception as e:
            print(f"❌ Error saving submission to Excel: {e}")
            traceback.print_exc()
            return False
    
    async def _update_existing_submission(self, data: dict, row: int) -> bool:
        """Update an existing submission in Excel"""
        try:
            wb = load_workbook(self.excel_file)
            ws = wb.active
            
            ws[f'A{row}'] = data.get('name', '').strip()
            ws[f'B{row}'] = data.get('role', '').strip()
            ws[f'C{row}'] = data.get('injectiveRole', '').strip()
            ws[f'D{row}'] = data.get('experience', '').strip()
            ws[f'E{row}'] = data.get('education', '').strip()
            ws[f'F{row}'] = data.get('location', '').strip()
            ws[f'G{row}'] = 'Yes' if data.get('available', False) else 'No'
            ws[f'H{row}'] = data.get('monthlyRate', '').strip()
            ws[f'I{row}'] = ', '.join(data.get('skills', []))
            ws[f'J{row}'] = ', '.join(data.get('languages', []))
            ws[f'K{row}'] = data.get('discord', '').strip()
            ws[f'L{row}'] = data.get('email', '').strip()
            ws[f'M{row}'] = data.get('phone', '').strip() or '-'
            ws[f'N{row}'] = data.get('telegram', '').strip() or '-'
            ws[f'O{row}'] = data.get('X', '').strip() or '-'
            ws[f'P{row}'] = data.get('github', '').strip() or '-'
            ws[f'R{row}'] = data.get('walletType', '').strip()
            ws[f'S{row}'] = data.get('nftHold', '').strip()
            ws[f'T{row}'] = data.get('tokenHold', '').strip()
            ws[f'U{row}'] = data.get('portfolio', '').strip() or '-'
            ws[f'V{row}'] = data.get('cv', '').strip()
            ws[f'W{row}'] = data.get('profilePicture', '').strip()
            ws[f'X{row}'] = data.get('bio', '').strip()
            ws[f'Y{row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            ws[f'Z{row}'] = "Pending"
            
            wb.save(self.excel_file)
            wb.close()
            print(f"✅ Updated existing submission in Excel for row: {row} (Status: Pending)")
            return True
            
        except Exception as e:
            print(f"❌ Error updating existing submission: {e}")
            traceback.print_exc()
            return False
    
    async def _test_excel_access(self):
        """Test if we can read/write to Excel file"""
        try:
            # Test reading
            wb = load_workbook(self.excel_file, read_only=True)
            ws = wb.active
            print(f"📊 Excel file has {ws.max_row} rows")
            wb.close()
            
            # Test writing
            wb = load_workbook(self.excel_file)
            ws = wb.active
            # Add a test cell
            ws['AA1'] = 'Test Write'
            wb.save(self.excel_file)
            wb.close()
            
            print("✅ Excel file is accessible for reading and writing")
            return True
        except Exception as e:
            print(f"❌ Excel access error: {e}")
            return False
    
    async def _process_new_submission(self, data: dict):
        """Process a new submission - save to Excel and notify"""
        try:
            wallet = data.get('walletAddress', '').strip()
            
            if not wallet:
                print(f"[SUBMISSION ERROR] No wallet address in data")
                return
            
            print(f"📢 Processing Discord notification for: {wallet}")
            
            # Send notification to Discord channel
            if self.bot and self._bot_running:
                channel = self.bot.get_channel(self.submission_channel_id)
                if channel:
                    # Create a more detailed notification embed
                    notification_embed = discord.Embed(
                        title="📬 **NEW SUBMISSION RECEIVED** 📬",
                        description=f"**{data.get('name', 'Unknown')}** has submitted a new talent profile.",
                        color=discord.Color.gold()
                    )
                    
                    # Add important fields
                    notification_embed.add_field(name="👤 Name", value=data.get('name', 'N/A'), inline=True)
                    notification_embed.add_field(name="💼 Role", value=data.get('role', 'N/A'), inline=True)
                    notification_embed.add_field(name="📍 Location", value=data.get('location', 'N/A'), inline=True)
                    
                    # Add experience and availability
                    notification_embed.add_field(name="📊 Experience", value=data.get('experience', 'N/A'), inline=True)
                    notification_embed.add_field(name="✅ Available", 
                                               value='Yes' if data.get('available', False) else 'No', 
                                               inline=True)
                    
                    # Add skills (truncated if too long)
                    skills = ', '.join(data.get('skills', []))
                    if len(skills) > 100:
                        skills = skills[:97] + '...'
                    notification_embed.add_field(name="🛠️ Skills", value=skills or 'N/A', inline=False)
                    
                    # Add contact info
                    contact_info = []
                    if data.get('discord'):
                        contact_info.append(f"**Discord:** {data.get('discord')}")
                    if data.get('email'):
                        contact_info.append(f"**Email:** {data.get('email')}")
                    
                    if contact_info:
                        notification_embed.add_field(name="📞 Contact Info", value='\n'.join(contact_info), inline=False)
                    
                    # Add wallet info
                    notification_embed.add_field(
                        name="💰 Wallet Address", 
                        value=f"`{wallet}`", 
                        inline=False
                    )
                    
                    # Add action buttons
                    view = discord.ui.View(timeout=None)
                    
                    # Review button
                    review_button = discord.ui.Button(
                        label="📝 Review Submission",
                        style=discord.ButtonStyle.primary,
                        custom_id=f"review:{wallet}"
                    )
                    
                    # Quick approve button
                    approve_button = discord.ui.Button(
                        label="✅ Quick Approve",
                        style=discord.ButtonStyle.success,
                        custom_id=f"quick_approve:{wallet}"
                    )
                    
                    # Add buttons to view
                    view.add_item(review_button)
                    view.add_item(approve_button)
                    
                    # Send ping for attention
                    await channel.send(f"📬 @here **New submission pending review!** 📬")
                    await channel.send(embed=notification_embed, view=view)
                    
                    # Also log to console
                    print(f"📢 Notification sent to Discord channel #{channel.name}")
                    
                else:
                    print(f"❌ Cannot find channel with ID {self.submission_channel_id}")
            else:
                print(f"⚠️ Bot not ready or not running, notification skipped")
                
        except Exception as e:
            print(f"[SUBMISSION ERROR] Error processing submission: {e}")
            traceback.print_exc()
    
    async def _handle_submission_interaction(self, interaction: discord.Interaction, action: str, wallet: str):
        """Handle submission button interactions"""
        print(f"[SUBMISSION INTERACTION] Processing {action} for wallet: {wallet}")
        
        try:
            success = False
            if action == "approve":
                success = await self._update_excel_status(wallet, "Approved")
                
                if success:
                    await interaction.followup.send(
                        f"✅ Submission approved for wallet `{wallet}`!",
                        ephemeral=True
                    )
                else:
                    await interaction.followup.send(
                        f"❌ Failed to approve submission for wallet `{wallet}`",
                        ephemeral=True
                    )
                    
            elif action == "reject":
                success = await self._update_excel_status(wallet, "Rejected")
                
                if success:
                    await interaction.followup.send(
                        f"❌ Submission rejected for wallet `{wallet}`!",
                        ephemeral=True
                    )
                else:
                    await interaction.followup.send(
                        f"❌ Failed to reject submission for wallet `{wallet}`",
                        ephemeral=True
                    )
                    
            elif action == "close":
                await interaction.followup.send(
                    f"🔒 Submission review closed for wallet `{wallet}`",
                    ephemeral=True
                )
            else:
                await interaction.followup.send(
                    f"❌ Unknown action: {action}",
                    ephemeral=True
                )
                return
            
            try:
                channel = self.bot.get_channel(self.submission_channel_id)
                if channel and success and action in ["approve", "reject"]:
                    status_embed = discord.Embed(
                        title=f"Submission {action.capitalize()}d",
                        description=f"Wallet: `{wallet}` has been **{action}d**.",
                        color=discord.Color.green() if action == "approve" else discord.Color.red()
                    )
                    status_embed.add_field(name="Action", value=f"**{action.upper()}**", inline=True)
                    status_embed.add_field(name="By", value=interaction.user.mention, inline=True)
                    status_embed.set_footer(text=f"Status updated in Excel to: {action.capitalize()}d")
                    
                    await channel.send(embed=status_embed)
            except Exception as e:
                print(f"[INTERACTION] Error sending status update: {e}")
                
        except Exception as e:
            print(f"[INTERACTION BACKGROUND ERROR] Error in background processing: {e}")
            traceback.print_exc()
            try:
                await interaction.followup.send(
                    "❌ An error occurred while processing your request!",
                    ephemeral=True
                )
            except:
                pass
    
    def submit_from_thread(self, data: dict) -> bool:
        """
        Thread-safe method to submit data from any thread.
        Returns True if submission was added to queue, False otherwise.
        """
        try:
            wallet = data.get('walletAddress', '').strip()
            print(f"[THREAD] Processing submission for: {wallet}")
            print(f"[THREAD] Submission details:")
            print(f"  - Name: {data.get('name')}")
            print(f"  - Role: {data.get('role')}")
            print(f"  - Wallet: {wallet}")
            
            # Save to Excel immediately (this doesn't need the bot to be ready)
            excel_success = False
            try:
                # Run the async save function in a new event loop
                loop = asyncio.new_event_loop()
                asyncio.set_event_loop(loop)
                excel_success = loop.run_until_complete(self._save_new_submission(data))
                loop.close()
            except Exception as e:
                print(f"[THREAD ERROR] Error saving to Excel: {e}")
                traceback.print_exc()
            
            if not excel_success:
                print(f"⚠️ [THREAD] Failed to save submission for {wallet} to Excel")
                return False
            
            print(f"✅ [THREAD] Saved to Excel: {wallet}")
            
            # Queue for Discord notification
            if self.bot and self._bot_running:
                print(f"[THREAD] Bot is running, scheduling Discord notification")
                # Schedule processing in bot's event loop
                if self._bot_loop:
                    try:
                        asyncio.run_coroutine_threadsafe(
                            self._process_new_submission(data),
                            self._bot_loop
                        )
                        print(f"[THREAD] Discord notification scheduled")
                    except Exception as e:
                        print(f"[THREAD ERROR] Failed to schedule notification: {e}")
                        traceback.print_exc()
                        # If scheduling fails, add to queue
                        self._thread_submission_queue.put(data)
                else:
                    print(f"⚠️ [THREAD] Bot loop not available, adding to queue")
                    self._thread_submission_queue.put(data)
            else:
                print(f"⚠️ [THREAD] Bot not ready yet, adding to queue")
                self._thread_submission_queue.put(data)
            
            return True
            
        except Exception as e:
            print(f"[THREAD ERROR] Failed to process submission: {e}")
            traceback.print_exc()
            return False
    
    async def _periodic_notifications(self):
        """Send periodic notifications about pending submissions"""
        print("⏰ Starting periodic notification system...")
        
        # Wait for bot to be fully ready
        await self.bot.wait_until_ready()
        
        while not self.bot.is_closed():
            try:
                # Wait for 1 hour (3600 seconds) between checks
                await asyncio.sleep(3600)
                
                # Get pending submissions count
                pending_records = await self._get_pending_records()
                
                if pending_records:
                    channel = self.bot.get_channel(self.submission_channel_id)
                    if channel:
                        # Don't send notification if there's already a recent one
                        # Check last 2 hours of messages
                        recent_notifications = False
                        async for message in channel.history(limit=50, after=discord.utils.utcnow() - timedelta(hours=2)):
                            if "pending submission" in message.content.lower() or "new submission" in message.content.lower():
                                recent_notifications = True
                                break
                        
                        if not recent_notifications and len(pending_records) > 0:
                            # Send reminder notification
                            reminder_embed = discord.Embed(
                                title="⏰ **PENDING SUBMISSIONS REMINDER** ⏰",
                                description=f"There are **{len(pending_records)}** submission(s) waiting for review.",
                                color=discord.Color.orange()
                            )
                            
                            # List first 5 pending submissions
                            submissions_list = []
                            for i, (name, wallet) in enumerate(pending_records[:5], 1):
                                submissions_list.append(f"{i}. **{name}** (`{wallet[:8]}...{wallet[-6:]}`)")
                            
                            if submissions_list:
                                reminder_embed.add_field(
                                    name="Pending Submissions",
                                    value="\n".join(submissions_list),
                                    inline=False
                                )
                            
                            reminder_embed.add_field(
                                name="Quick Actions",
                                value="• Use `/job_open` to review submissions\n• Use `/job_status` to see all status counts",
                                inline=False
                            )
                            
                            reminder_embed.set_footer(text="Last checked: " + datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
                            
                            await channel.send(f"📬 **Reminder:** {len(pending_records)} pending submission(s) need review! 📬")
                            await channel.send(embed=reminder_embed)
                            print(f"⏰ Sent periodic reminder for {len(pending_records)} pending submissions")
                            
            except Exception as e:
                print(f"[PERIODIC NOTIFICATION ERROR] {e}")
                traceback.print_exc()

    def start_bot(self):
        """Start the Discord bot (call this separately, not in Django's main thread)"""
        if not self.bot_code:
            print("❌ Cannot start bot: No token provided!")
            return False
            
        print("🤖 Starting Discord Bot...")
        print("=" * 50)
        print("WORKFLOW:")
        print("1. New submissions are saved to Excel with status 'Pending'")
        print("2. Bot sends notification when new submission arrives")
        print("3. Use /job_open to see and review pending submissions")
        print("4. Use /job_status to see counts by status")
        print("5. Use /job_show to see approved submissions")
        print("=" * 50)
        
        try:
            if self._bot_running:
                print("✅ Bot is already running")
                return True
                
            self.bot.run(self.bot_code)
            return True
        except discord.LoginFailure:
            print("❌ Failed to login. Check your bot token!")
            return False
        except RuntimeError as e:
            if "asyncio.run() cannot be called from a running event loop" in str(e):
                print("⚠️ Bot already running in another event loop")
                return True
            print(f"❌ Error starting bot: {e}")
            traceback.print_exc()
            return False
        except Exception as e:
            print(f"❌ Error starting bot: {e}")
            traceback.print_exc()
            return False

    def start_bot_async(self):
        """Start the Discord bot in a separate thread for Django compatibility"""
        if TalentHubBot._bot_started:
            print("✅ Discord bot already started")
            return None
            
        import threading
        
        def run_bot():
            try:
                loop = asyncio.new_event_loop()
                asyncio.set_event_loop(loop)
                
                if self._bot_running:
                    print("✅ Bot is already running")
                    return
                
                print("🤖 Starting Discord bot in separate thread...")
                loop.run_until_complete(self.bot.start(self.bot_code))
            except Exception as e:
                print(f"❌ Error in bot thread: {e}")
                traceback.print_exc()
        
        if not hasattr(self, '_bot_thread') or not self._bot_thread.is_alive():
            self._bot_thread = threading.Thread(target=run_bot, daemon=True, name="DiscordBotThread")
            self._bot_thread.start()
            TalentHubBot._bot_started = True
            print("🤖 Discord bot thread started")
        else:
            print("✅ Discord bot thread is already running")
        
        return self._bot_thread


talent_hub_bot = TalentHubBot()